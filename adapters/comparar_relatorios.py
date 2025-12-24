"""
Script para comparar relatórios WMS (C#) vs API (Python) e gerar Excel com comparação lado a lado.
"""
import re
from pathlib import Path
from typing import List, Dict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# Variável global para rastrear arquivos processados: {number_mapa: (wms_file, api_file)}
_MAPEAMENTO_ARQUIVOS = {}

# Paleta de cores global - usada em TODOS os lugares para garantir consistência
PALETA_CORES_PALLETS = [
    "E0E0E0",  # Cinza médio claro
    "CCE5FF",  # Azul claro perceptível
    "FFE4CC",  # Laranja bem claro
    "CCFFCC",  # Verde claro perceptível
    "FFCCFF",  # Rosa claro perceptível
    "E6E6B8",  # Amarelo suave
    "D4E6F1",  # Azul acinzentado
    "D5F4E6",  # Verde menta perceptível
    "FADBD8",  # Salmão claro
    "E8DAEF",  # Lavanda claro
]


def padronizar_pallet_code(pallet_code: str) -> str:
    """
    Padroniza formato do pallet_code para P03_A_03_1/35 (com zeros).
    O número após P deve ser o mesmo número que vem após A_ ou M_.
    
    Exemplos:
        P_A_01_1/35  -> P01_A_01_1/35
        P01_M_1_1/35 -> P01_M_01_1/35
        P_M_02_1/35  -> P02_M_02_1/35
        P_A_03_1/35  -> P03_A_03_1/35
    """
    # Pattern: P(número?)_(A|M)_(número)_(número/número)
    match = re.match(r'P(\d*)_([AM])_(\d+)_(\d+/\d+)', pallet_code)
    if match:
        num1 = match.group(1)  # Número após P (pode estar vazio)
        lado = match.group(2)  # A ou M
        num2 = match.group(3)  # Número após A_ ou M_
        final = match.group(4)  # Parte final (ex: 1/35)
        
        # Se num1 estiver vazio, usa num2 como referência
        # Se num1 existir, usa ele mesmo
        num_pallet = num2 if not num1 else num1
        
        # Formata com zeros à esquerda (2 dígitos)
        num_pallet_pad = num_pallet.zfill(2)
        num2_pad = num2.zfill(2)
        
        return f"P{num_pallet_pad}_{lado}_{num2_pad}_{final}"
    
    return pallet_code


def extrair_produtos_do_relatorio(filepath: Path) -> List[Dict]:
    """
    Extrai produtos de um arquivo de relatório TXT.
    
    Abordagem híbrida: usa regex flexível que não depende de tamanhos fixos.
    Captura grupos principais e extrai campos de forma adaptativa.
    
    Retorna lista de dicionários com:
    - number_mapa: código do mapa (ex: 620807)
    - pallet_code: código do pallet (ex: P01_A_01_1/35)
    - product_code: código do produto
    - product_name: nome do produto
    - quantity: quantidade
    - atributo: atributo do produto
    - ocupacao: ocupação
    """
    produtos = []
    
    with open(filepath, 'r', encoding='utf-8') as f:
        conteudo = f.read()
    
    # Extrai número do mapa da primeira linha (formato: "Mapa: 620807 Veículo: ...")
    number_mapa = None
    for linha in conteudo.split('\n'):
        match_mapa = re.match(r'Mapa:\s*(\d+)', linha)
        if match_mapa:
            number_mapa = match_mapa.group(1)
            break
    
    # Pattern para capturar o código do pallet
    pallet_pattern = r'^(P\d*_[AM]_\d+_\d+/\d+)\s+-\s+'
    
    pallet_atual = None
    
    for linha in conteudo.split('\n'):
        # Verifica se é uma linha de cabeçalho de pallet
        match_pallet = re.match(pallet_pattern, linha.strip())
        if match_pallet:
            pallet_atual = match_pallet.group(1)
            # Padroniza formato: P01_M_01_1/35 (adiciona zero se necessário)
            pallet_atual = padronizar_pallet_code(pallet_atual)
            continue
        
        # Verifica se é uma linha de produto válida
        if pallet_atual and linha.strip().startswith('|') and linha.strip().endswith('|'):
            # Ignora separadores
            if '====' in linha:
                continue
            
            # Remove pipes das pontas
            linha_limpa = linha.strip()[1:-1].strip()
            
            # Ignora linhas vazias
            if not linha_limpa:
                continue
            
            # Pattern flexível: captura os componentes principais
            # Estrutura: | index code nome_produto quantidade embalagem grupo/sub peso atributo ocupacao |
            # Regex mais flexível: captura code (dígitos), depois texto até encontrar padrão de quantidade
            match = re.match(
                r'\s*\d+\s+'           # index (sempre 0)
                r'(\d+)\s+'            # product_code (grupo 1)
                r'(.+?)\s+'            # product_name (grupo 2) - non-greedy
                r'(\d+[A-Z]?)\s+'      # quantity (grupo 3) - pode ter letra no final
                r'\d{4}\s+'            # embalagem (não captura)
                r'\d+/\d+\s+'          # grupo/sub (não captura)
                r'[\d.,]+\s+'          # peso (não captura)
                r'(.+?)\s*'            # atributo (grupo 4)
                r'([\d.,]*)\s*$',      # ocupacao (grupo 5) - OPCIONAL (pode estar vazia)
                linha_limpa
            )
            
            if match:
                product_code = match.group(1)
                product_name = match.group(2).strip()
                quantity = match.group(3)
                atributo = match.group(4)
                ocupacao = match.group(5).replace(',', '.')
                
                # Remove 'A' do final da quantidade se existir (indica avulso)
                quantity_clean = quantity.rstrip('A')
                
                try:
                    produtos.append({
                        'number_mapa': number_mapa,
                        'pallet_code': pallet_atual,
                        'product_code': product_code,
                        'product_name': product_name,
                        'quantity': float(quantity_clean),
                        'atributo': atributo,
                        'ocupacao': float(ocupacao) if ocupacao else 0.0
                    })
                except ValueError:
                    # Ignora se não conseguir converter
                    pass
    
    return produtos


def extrair_produtos_nao_paletizados(filepath: Path, number_mapa: str = None) -> List[Dict]:
    """
    Extrai produtos da seção 'Produtos fora do caminhão' do relatório TXT.
    
    Args:
        filepath: Caminho para o arquivo TXT
        number_mapa: Número do mapa (opcional, será extraído do arquivo se não fornecido)
    
    Retorna lista de dicionários com:
    - number_mapa: código do mapa
    - product_code: código do produto
    - product_name: nome do produto
    - quantity: quantidade
    - atributo: atributo do produto
    - ocupacao: ocupação
    """
    produtos = []
    
    print(f"\n[INFO] Extraindo produtos nao paletizados de: {filepath}")
    
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            conteudo = f.read()
    except Exception as e:
        print(f"[ERRO] Erro ao ler arquivo: {e}")
        return produtos
    
    # Usa number_mapa fornecido ou tenta extrair do arquivo
    if number_mapa is None:
        for linha in conteudo.split('\n'):
            match_mapa = re.match(r'Mapa:\s*(\d+)', linha)
            if match_mapa:
                number_mapa = match_mapa.group(1)
                break
            break
    
    print(f"   Number_mapa encontrado: {number_mapa}")
    
    # Procura pela seção "Produtos fora do caminhão"
    linhas = conteudo.split('\n')
    dentro_secao = False
    linhas_processadas = 0
    
    for i, linha in enumerate(linhas):
        # Identifica início da seção
        if 'Produtos fora do caminhão' in linha:
            dentro_secao = True
            print(f"   [OK] Secao 'Produtos fora do caminhao' encontrada na linha {i}")
            continue
        
        # Se ainda não entrou na seção, pula
        if not dentro_secao:
            continue
        
        linhas_processadas += 1
        
        # Para na próxima seção vazia ou quando encontrar linha sem conteúdo relevante
        if linhas_processadas > 100:  # Limite de segurança
            break
        
        # Verifica se é linha de produto (formato similar aos pallets)
        if linha.strip().startswith('|') and linha.strip().endswith('|'):
            # Ignora separadores
            if '====' in linha or '----' in linha:
                continue
            
            # Remove pipes das pontas
            linha_limpa = linha.strip()[1:-1].strip()
            
            # Ignora linhas vazias
            if not linha_limpa:
                continue
            
            print(f"   Tentando parsear linha: '{linha_limpa[:80]}...'")
            
            # Pattern para produtos não paletizados (SEM índice, ocupacao OPCIONAL)
            # Formato WMS: | product_code product_name quantity embalagem grupo/sub peso atributo ocupacao |
            # Formato API: | product_code product_name quantity embalagem grupo/sub peso atributo |
            match = re.match(
                r'\s*(\d+)\s+'         # product_code (grupo 1)
                r'(.+?)\s+'            # product_name (grupo 2)
                r'(\d+[A-Z]?)\s+'      # quantity (grupo 3)
                r'\d{4}\s+'            # embalagem
                r'\d+/\d+\s+'          # grupo/sub
                r'([\d.,]+)\s+'        # peso (grupo 4)
                r'(\S+)\s*'            # atributo (grupo 5)
                r'([\d.,]*)\s*$',      # ocupacao OPCIONAL (grupo 6)
                linha_limpa
            )
            
            if match:
                product_code = match.group(1)
                product_name = match.group(2).strip()
                quantity = match.group(3)
                peso = match.group(4).replace(',', '.')  # Normaliza decimal
                atributo = match.group(5)
                ocupacao = match.group(6).replace(',', '.') if match.group(6) else ''
                
                quantity_clean = quantity.rstrip('A')
                
                # Calcula ocupacao se não vier no arquivo (caso API)
                if not ocupacao:
                    try:
                        # Ocupacao = peso / 30.75 (fórmula padrão)
                        ocupacao_calculada = float(peso) / 30.75
                        ocupacao = str(round(ocupacao_calculada, 2))
                        print(f"   [CALC] Ocupacao calculada: {ocupacao} (peso={peso})")
                    except:
                        ocupacao = '0.0'
                
                try:
                    produtos.append({
                        'number_mapa': number_mapa,
                        'product_code': product_code,
                        'product_name': product_name,
                        'quantity': float(quantity_clean),
                        'atributo': atributo,
                        'ocupacao': float(ocupacao) if ocupacao else 0.0
                    })
                    print(f"   [OK] Produto extraido: {product_code} - {product_name[:30]}, ocupacao={ocupacao}")
                except ValueError as e:
                    print(f"   [ERRO] Erro ao converter valores: {e}")
            else:
                print(f"   [WARN] Linha nao deu match no regex")
    
    print(f"   [INFO] Total de produtos nao paletizados extraidos: {len(produtos)}\n")
    return produtos


def criar_excel_comparacao(wms_file: Path, api_file: Path, output_file: Path, append: bool = False):
    """
    Cria arquivo Excel com comparação lado a lado entre WMS e API.
    
    Args:
        wms_file: Arquivo TXT do WMS
        api_file: Arquivo TXT da API
        output_file: Arquivo Excel de saída
        append: Se True, concatena com dados existentes na planilha
    """
    print(f"Lendo arquivo WMS: {wms_file}")
    produtos_wms = extrair_produtos_do_relatorio(wms_file)
    
    print(f"Lendo arquivo API: {api_file}")
    produtos_api = extrair_produtos_do_relatorio(api_file)
    
    print(f"WMS: {len(produtos_wms)} produtos extraídos")
    print(f"API: {len(produtos_api)} produtos extraídos")
    
    # Extrai number_mapa (prioriza API, se não tiver usa WMS)
    number_mapa = None
    if produtos_api and produtos_api[0].get('number_mapa'):
        number_mapa = produtos_api[0]['number_mapa']
    elif produtos_wms and produtos_wms[0].get('number_mapa'):
        number_mapa = produtos_wms[0]['number_mapa']
    
    # Se ainda não tiver number_mapa, tenta extrair do nome do arquivo
    if not number_mapa:
        # Tenta extrair do nome do arquivo API (formato: palletize_result_map_XXXXXX.txt)
        import re
        match = re.search(r'map[_-]?(\d+)', str(api_file.name))
        if match:
            number_mapa = match.group(1)
            print(f"⚠️ number_mapa não encontrado nos arquivos, extraído do nome: {number_mapa}")
    
    print(f"Number_mapa detectado: {number_mapa}")
    
    # CRÍTICO: Aplica number_mapa em TODOS os produtos (WMS e API)
    if number_mapa:
        print(f"Aplicando number_mapa '{number_mapa}' em todos os produtos...")
        for produto in produtos_wms:
            produto['number_mapa'] = number_mapa
        for produto in produtos_api:
            produto['number_mapa'] = number_mapa
    else:
        print(f"⚠️ AVISO: number_mapa não pôde ser determinado!")
    
    # Cria DataFrames
    df_wms_new = pd.DataFrame(produtos_wms)
    df_api_new = pd.DataFrame(produtos_api)
    
    # Verifica se number_mapa foi aplicado corretamente
    if not df_wms_new.empty and df_wms_new['number_mapa'].isna().all():
        print(f"⚠️ AVISO: number_mapa não foi preenchido no WMS!")
    if not df_api_new.empty and df_api_new['number_mapa'].isna().all():
        print(f"⚠️ AVISO: number_mapa não foi preenchido na API!")
    
    # Garante que os DataFrames tenham as colunas esperadas mesmo se estiverem vazios
    colunas_esperadas = ['number_mapa', 'pallet_code', 'product_code', 'product_name', 'quantity', 'atributo', 'ocupacao']
    
    if df_wms_new.empty:
        print("⚠️ WMS novo está vazio, criando DataFrame com colunas padrão")
        df_wms_new = pd.DataFrame(columns=colunas_esperadas)
    
    if df_api_new.empty:
        print("⚠️ API novo está vazio, criando DataFrame com colunas padrão")
        df_api_new = pd.DataFrame(columns=colunas_esperadas)
    
    # Se append=True e arquivo existe, carrega dados existentes
    if append and output_file.exists():
        print(f"\n{'='*60}")
        print(f"[APPEND] MODO APPEND ATIVADO")
        print(f"{'='*60}")
        print(f"[INFO] Arquivo existente: {output_file}")
        try:
            # Carrega dados existentes (pula linha 11 de cabeçalho, linha 12 tem os nomes de coluna)
            df_existing = pd.read_excel(output_file, sheet_name='Comparação', header=11)
            
            print(f"[INFO] Dados carregados do arquivo:")
            print(f"   Total de linhas no DataFrame: {len(df_existing)}")
            print(f"   Total de colunas: {len(df_existing.columns)}")
            print(f"   Colunas encontradas (com índices):")
            for i, col in enumerate(df_existing.columns):
                print(f"      [{i}] = '{col}'")
            
            # CRÍTICO: Verifica se temos as 15 colunas esperadas
            if len(df_existing.columns) < 15:
                print(f"\n   [WARN] PROBLEMA: Esperado 15 colunas, encontrado apenas {len(df_existing.columns)}")
                print(f"   Isso vai causar erro na extração da API!")
            
            # Define nomes das colunas esperadas
            colunas_wms = ['number_mapa', 'pallet_code', 'product_code', 'product_name', 'quantity', 'atributo', 'ocupacao']
            colunas_api = ['number_mapa', 'pallet_code', 'product_code', 'product_name', 'quantity', 'atributo', 'ocupacao']
            
            # Extrai DataFrames WMS e API existentes usando índices fixos de colunas
            # WMS: colunas 0-6 (7 colunas)
            # Separador: coluna 7 (1 coluna)  
            # API: colunas 8-14 (7 colunas)
            print(f"\n[INFO] Extraindo dados antigos:")
            print(f"   WMS: colunas 0 a 6")
            print(f"   Separador: coluna 7")
            print(f"   API: colunas 8 a 14")
            
            df_wms_old = df_existing.iloc[:, 0:7].copy()
            
            # CRÍTICO: Verifica se temos colunas suficientes para API
            if len(df_existing.columns) >= 15:
                df_api_old = df_existing.iloc[:, 8:15].copy()
            else:
                print(f"\n   [ERRO] ERRO CRITICO: Nao ha colunas suficientes para extrair API!")
                print(f"   Criando DataFrame API vazio...")
                df_api_old = pd.DataFrame(columns=colunas_api)
            
            print(f"   WMS antigo: {df_wms_old.shape}")
            print(f"   API antigo: {df_api_old.shape}")
            
            # Renomeia colunas para garantir consistência
            df_wms_old.columns = colunas_wms
            df_api_old.columns = colunas_api
            
            # Remove linhas vazias - tanto NaN quanto strings vazias
            # Primeiro substitui strings vazias por NaN
            df_wms_old = df_wms_old.replace('', pd.NA)
            df_api_old = df_api_old.replace('', pd.NA)
            
            # Remove apenas linhas onde TODAS as colunas são NaN/vazias
            # IMPORTANTE: Não remove linhas que tenham pelo menos number_mapa preenchido
            # porque isso indica dados válidos mesmo que outras colunas estejam vazias
            
            # Para WMS: remove apenas se TODAS as colunas estão vazias
            df_wms_old = df_wms_old.dropna(how='all')
            
            # Para API: CRÍTICO - não remove linhas se number_mapa estiver preenchido
            # ou se pelo menos product_code estiver preenchido (indica dado real)
            # Remove apenas linhas completamente vazias OU com apenas espaços
            mask_api_valido = (
                df_api_old['number_mapa'].notna() | 
                df_api_old['product_code'].notna() |
                df_api_old['pallet_code'].notna()
            )
            
            # Se NENHUMA linha for válida, então o DataFrame está vazio
            if mask_api_valido.any():
                df_api_old = df_api_old[mask_api_valido]
            else:
                # Todas as linhas são inválidas, remove todas
                df_api_old = df_api_old.dropna(how='all')
            
            # Substitui NaN de volta para '' se necessário (para consistência)
            df_wms_old = df_wms_old.fillna('')
            df_api_old = df_api_old.fillna('')
            
            # Substitui NaN de volta para '' se necessário (para consistência)
            df_wms_old = df_wms_old.fillna('')
            df_api_old = df_api_old.fillna('')
            
            print(f"\n[INFO] Apos remover linhas vazias:")
            print(f"   WMS antigo: {len(df_wms_old)} linhas")
            print(f"   API antigo: {len(df_api_old)} linhas")
            print(f"   WMS novo: {len(df_wms_new)} linhas")
            print(f"   API novo: {len(df_api_new)} linhas")
            

            
            # CRÍTICO: Antes de concatenar, verifica se TODOS os DataFrames têm as mesmas colunas
            print(f"\n[INFO] VERIFICACAO DE COLUNAS ANTES DA CONCATENACAO:")
            print(f"   df_wms_old colunas: {list(df_wms_old.columns)}")
            print(f"   df_wms_new colunas: {list(df_wms_new.columns)}")
            print(f"   df_api_old colunas: {list(df_api_old.columns)}")
            print(f"   df_api_new colunas: {list(df_api_new.columns)}")
            
            # Garante que todos os DataFrames tenham EXATAMENTE as mesmas colunas ANTES da concatenação
            # Isso evita problemas como o que tivemos quando 'atributo' estava faltando
            for col in colunas_wms:
                if col not in df_wms_old.columns:
                    print(f"   [WARN] Adicionando coluna '{col}' faltante em df_wms_old")
                    df_wms_old[col] = ''
                if col not in df_wms_new.columns:
                    print(f"   ⚠️ Adicionando coluna '{col}' faltante em df_wms_new")
                    df_wms_new[col] = ''
                    
            for col in colunas_api:
                if col not in df_api_old.columns:
                    print(f"   ⚠️ Adicionando coluna '{col}' faltante em df_api_old")
                    df_api_old[col] = ''
                if col not in df_api_new.columns:
                    print(f"   ⚠️ Adicionando coluna '{col}' faltante em df_api_new")
                    df_api_new[col] = ''
            
            # Garante ordem correta das colunas ANTES da concatenação
            df_wms_old = df_wms_old[colunas_wms]
            df_wms_new = df_wms_new[colunas_wms]
            df_api_old = df_api_old[colunas_api]
            df_api_new = df_api_new[colunas_api]
            
            print(f"\n✅ Colunas padronizadas! Agora todos têm: {colunas_wms}")
            
            # Concatena verticalmente (novos dados embaixo dos antigos)
            print(f"\n🔗 Realizando concatenação vertical:")
            print(f"   df_wms_old: {len(df_wms_old)} linhas")
            print(f"   df_wms_new: {len(df_wms_new)} linhas")
            print(f"   df_api_old: {len(df_api_old)} linhas")
            print(f"   df_api_new: {len(df_api_new)} linhas")
            
            df_wms = pd.concat([df_wms_old, df_wms_new], ignore_index=True)
            df_api = pd.concat([df_api_old, df_api_new], ignore_index=True)
            
            print(f"\n🔗 Após concatenação vertical:")
            print(f"   df_wms: {len(df_wms)} linhas (esperado: {len(df_wms_old) + len(df_wms_new)})")
            print(f"   df_api: {len(df_api)} linhas (esperado: {len(df_api_old) + len(df_api_new)})")
            
            # CRÍTICO: Verifica se a concatenação funcionou
            if len(df_wms) != len(df_wms_old) + len(df_wms_new):
                print(f"   ⚠️ AVISO: WMS concatenação com problema!")
            if len(df_api) != len(df_api_old) + len(df_api_new):
                print(f"   ⚠️ AVISO: API concatenação com problema!")
            
            # Garante que as colunas estão na ordem correta
            df_wms = df_wms[colunas_wms]
            df_api = df_api[colunas_api]
            
            print(f"\n✅ CONCATENAÇÃO REALIZADA:")
            print(f"   ✓ Total WMS: {len(df_wms)} linhas ({len(df_wms_old)} antigas + {len(df_wms_new)} novas)")
            print(f"   ✓ Total API: {len(df_api)} linhas ({len(df_api_old)} antigas + {len(df_api_new)} novas)")
            
            # DEBUG: Mostra amostra após concatenação VERTICAL
            if len(df_wms) > 0:
                print(f"\n   📊 WMS concatenado - primeiras 3 linhas:")
                print(f"     {df_wms.head(3)[['number_mapa', 'pallet_code', 'product_code']].to_dict('records')}")
                print(f"   📊 WMS concatenado - últimas 3 linhas:")
                print(f"     {df_wms.tail(3)[['number_mapa', 'pallet_code', 'product_code']].to_dict('records')}")
            
            if len(df_api) > 0:
                print(f"\n   📊 API concatenado - primeiras 3 linhas:")
                print(f"     {df_api.head(3)[['number_mapa', 'pallet_code', 'product_code']].to_dict('records')}")
                print(f"   📊 API concatenado - últimas 3 linhas:")
                print(f"     {df_api.tail(3)[['number_mapa', 'pallet_code', 'product_code']].to_dict('records')}")
            else:
                print(f"\n   ❌ CRÍTICO: df_api está VAZIO após concatenação vertical!")
            
            print(f"{'='*60}\n")
        except Exception as e:
            print(f"\n❌ ERRO ao carregar arquivo existente: {e}")
            import traceback
            traceback.print_exc()
            print(f"⚠️ Criando novo arquivo ao invés de concatenar...")
            df_wms = df_wms_new
            df_api = df_api_new
            append = False  # Força criação de novo arquivo
    else:
        print(f"\n{'='*60}")
        print(f"[CRIACAO] MODO CRIACAO - Novo arquivo")
        print(f"{'='*60}")
        print(f"   WMS: {len(df_wms_new)} linhas")
        print(f"   API: {len(df_api_new)} linhas")
        df_wms = df_wms_new
        df_api = df_api_new
    
    # Registra arquivos no mapeamento global para uso posterior (aba Não Paletizados)
    # CRÍTICO: Usa df_wms_new (novos dados) ao invés de df_wms (concatenado)
    global _MAPEAMENTO_ARQUIVOS
    if not df_wms_new.empty and 'number_mapa' in df_wms_new.columns:
        number_mapa = str(df_wms_new['number_mapa'].iloc[0])
        _MAPEAMENTO_ARQUIVOS[number_mapa] = (wms_file, api_file)
        print(f"[DEBUG] Mapeamento registrado: mapa {number_mapa} -> ({wms_file.name}, {api_file.name})")
    else:
        print(f"[WARN] Não foi possível registrar mapeamento (df_wms_new vazio ou sem number_mapa)")
    
    # Adiciona coluna vazia para separação
    max_linhas = max(len(df_wms), len(df_api))
    
    print(f"\n[INFO] Preparando DataFrame final:")
    print(f"   df_wms shape: {df_wms.shape}")
    print(f"   df_api shape: {df_api.shape}")
    print(f"   max_linhas: {max_linhas}")
    
    # Garante que ambos os DataFrames tenham o mesmo número de linhas
    # Preenche com linhas vazias se necessário
    if len(df_wms) < max_linhas:
        linhas_faltando = max_linhas - len(df_wms)
        print(f"   ⚠️ WMS tem menos linhas, adicionando {linhas_faltando} linhas vazias")
        df_wms_vazio = pd.DataFrame('', index=range(linhas_faltando), columns=df_wms.columns)
        # CRÍTICO: Copia number_mapa do df_api correspondente para as linhas vazias do WMS
        if len(df_api) > len(df_wms):
            for i in range(linhas_faltando):
                idx_api = len(df_wms) + i
                if idx_api < len(df_api) and 'number_mapa' in df_api.columns:
                    df_wms_vazio.loc[i, 'number_mapa'] = df_api.iloc[idx_api]['number_mapa']
        df_wms = pd.concat([df_wms, df_wms_vazio], ignore_index=True)
    
    if len(df_api) < max_linhas:
        linhas_faltando = max_linhas - len(df_api)
        print(f"   ⚠️ API tem menos linhas, adicionando {linhas_faltando} linhas vazias")
        df_api_vazio = pd.DataFrame('', index=range(linhas_faltando), columns=df_api.columns)
        # CRÍTICO: Copia number_mapa do df_wms correspondente para as linhas vazias da API
        # Isso garante que quando recarregarmos o arquivo, essas linhas não sejam removidas
        if len(df_wms) > len(df_api):
            for i in range(linhas_faltando):
                idx_wms = len(df_api) + i
                if idx_wms < len(df_wms) and 'number_mapa' in df_wms.columns:
                    df_api_vazio.loc[i, 'number_mapa'] = df_wms.iloc[idx_wms]['number_mapa']
        df_api = pd.concat([df_api, df_api_vazio], ignore_index=True)
    
    # Reseta os índices para garantir alinhamento correto
    df_wms = df_wms.reset_index(drop=True)
    df_api = df_api.reset_index(drop=True)
    
    # LIMPEZA: Remove espaços extras de todas as colunas de texto
    print(f"\n[INFO] Limpando espacos extras dos dados...")
    colunas_texto = ['number_mapa', 'pallet_code', 'product_code', 'product_name', 'atributo']
    for col in colunas_texto:
        if col in df_wms.columns:
            df_wms[col] = df_wms[col].astype(str).str.strip()
        if col in df_api.columns:
            df_api[col] = df_api[col].astype(str).str.strip()
    
    # Cria separador com o número correto de linhas
    df_separador = pd.DataFrame({'': [''] * max_linhas})
    
    print(f"   Após ajuste:")
    print(f"   df_wms shape: {df_wms.shape}")
    print(f"   df_api shape: {df_api.shape}")
    print(f"   df_separador shape: {df_separador.shape}")
    
    # DEBUG: Verifica df_api ANTES da concatenação horizontal
    print(f"\n[DEBUG] df_api ANTES da concatenacao horizontal:")
    print(f"   Shape: {df_api.shape}")
    if len(df_api) > 0:
        print(f"   Primeiras 20 linhas (pallet_code, product_code):")
        for i in range(min(20, len(df_api))):
            pallet = df_api.iloc[i]['pallet_code'] if 'pallet_code' in df_api.columns else 'N/A'
            product = df_api.iloc[i]['product_code'] if 'product_code' in df_api.columns else 'N/A'
            print(f"     Linha {i}: pallet={pallet}, product={product}")
    
    # Concatena horizontalmente: WMS + separador + API
    df_final = pd.concat([df_wms, df_separador, df_api], axis=1, ignore_index=False)
    
    # Preenche células vazias com ''
    df_final = df_final.fillna('')
    
    # Validação: Verifica se o DataFrame final tem as 15 colunas esperadas
    # 7 WMS + 1 separador + 7 API = 15 colunas
    colunas_esperadas = 15
    if len(df_final.columns) != colunas_esperadas:
        print(f"\n⚠️ AVISO: DataFrame final tem {len(df_final.columns)} colunas, esperado {colunas_esperadas}")
        print(f"   Colunas: {list(df_final.columns)}")
    else:
        print(f"\n✅ DataFrame final tem {len(df_final.columns)} colunas (correto!)")
    
    # DEBUG: Verifica se há dados da API
    print(f"\n🔍 Verificação da API no DataFrame final:")
    if len(df_final.columns) > 8:
        print(f"   Coluna 8 (index 8 - number_mapa API): {df_final.columns[8]}")
        print(f"   Coluna 9 (index 9 - pallet_code API): {df_final.columns[9]}")
        print(f"   Coluna 10 (index 10 - product_code API): {df_final.columns[10]}")
        
        # Mostra primeiras 20 linhas completas do lado API
        print(f"\n   📊 Primeiras 20 linhas do lado API (colunas 8-14):")
        for i in range(min(20, len(df_final))):
            row_api = df_final.iloc[i, 8:15].tolist()
            print(f"   Linha {i}: {row_api}")
        
        # Conta quantas linhas têm dados na API
        linhas_com_dados_api = (df_final.iloc[:, 8] != '').sum()
        print(f"\n   Linhas com dados na coluna number_mapa API: {linhas_com_dados_api} de {len(df_final)}")
        
        # Conta quantas têm pallet_code preenchido
        linhas_com_pallet_api = (df_final.iloc[:, 9] != '').sum()
        print(f"   Linhas com pallet_code API preenchido: {linhas_com_pallet_api} de {len(df_final)}")
    else:
        print(f"   ❌ ERRO: DataFrame não tem colunas suficientes para API!")
    
    # Coleta atributos únicos e pallets únicos para a tabela dinâmica
    # Verifica se os DataFrames têm dados antes de acessar colunas
    if len(df_wms) > 0 and 'atributo' in df_wms.columns:
        atributos_wms = sorted(df_wms['atributo'].unique().tolist())
        # IMPORTANTE: Mantém a ordem de aparição dos pallets (não ordena alfabeticamente)
        # Isso garante que as cores sigam a sequência da tabela WMS
        # Filtra apenas pallets não vazios
        pallets_wms = df_wms[df_wms['pallet_code'].notna() & (df_wms['pallet_code'] != '')]['pallet_code'].drop_duplicates().tolist()
    else:
        print("⚠️ DataFrame WMS vazio ou sem coluna 'atributo'")
        atributos_wms = []
        pallets_wms = []
    
    if len(df_api) > 0 and 'atributo' in df_api.columns:
        atributos_api = sorted(df_api['atributo'].unique().tolist())
        # IMPORTANTE: Mantém a ordem de aparição dos pallets (não ordena alfabeticamente)
        # Isso garante que as cores sigam a sequência da tabela API
        # CRÍTICO: Filtra apenas pallets não vazios (exclui linhas só com number_mapa)
        pallets_api = df_api[df_api['pallet_code'].notna() & (df_api['pallet_code'] != '')]['pallet_code'].drop_duplicates().tolist()
    else:
        print("⚠️ DataFrame API vazio ou sem coluna 'atributo'")
        atributos_api = []
        pallets_api = []
    
    # Salva no Excel
    print(f"\n💾 Salvando no Excel: {output_file}")
    print(f"   DataFrame final:")
    print(f"   - Shape: {df_final.shape}")
    print(f"   - Linhas: {len(df_final)}")
    print(f"   - Colunas: {len(df_final.columns)}")
    print(f"   - Colunas WMS (0-6): {list(df_final.columns[0:7])}")
    print(f"   - Coluna Separador (7): {df_final.columns[7]}")
    print(f"   - Colunas API (8-14): {list(df_final.columns[8:15])}")
    
    # DEBUG: Mostra amostra dos dados que serão salvos
    print(f"\n   📄 Amostra do DataFrame final (primeiras 3 linhas):")
    print(f"   WMS number_mapa: {df_final.iloc[:3, 0].tolist()}")
    print(f"   WMS pallet_code: {df_final.iloc[:3, 1].tolist()}")
    print(f"   API number_mapa: {df_final.iloc[:3, 8].tolist()}")
    print(f"   API pallet_code: {df_final.iloc[:3, 9].tolist()}")
    
    # IMPORTANTE: mode='w' sobrescreve completamente o arquivo
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
        df_final.to_excel(writer, sheet_name='Comparação', index=False, startrow=11)
    
    print(f"   ✓ Dados salvos no Excel")
    
    # Formata o Excel e cria tabela dinâmica
    print(f"\n🎨 Aplicando formatação...")
    # Em modo append (batch), não cria aba Não Paletizados ainda (será criada no final)
    formatar_excel(output_file, atributos_wms, atributos_api, pallets_wms, pallets_api, 
                   is_append=append, wms_file=wms_file, api_file=api_file, 
                   criar_aba_nao_paletizados_flag=not append)  # Só cria se NÃO for append
    
    print(f"\n✅ Arquivo criado com sucesso: {output_file}")
    print(f"{'='*60}\n")


def criar_tabela_dinamica_com_formulas(ws, nome_tabela_origem, inicio_linha, inicio_col, pallets, atributos, is_wms=True, range_dados=None, mapa_cores_pallets=None):
    """
    Cria tabela dinâmica com fórmulas SUMIFS que referenciam a tabela principal.
    
    Args:
        ws: Worksheet
        nome_tabela_origem: Nome da tabela do Excel (ex: 'TabelaWMS')
        inicio_linha: Linha onde começar a tabela pivot
        inicio_col: Coluna onde começar (índice baseado em 1)
        pallets: Lista de pallet_codes únicos
        atributos: Lista de atributos únicos
        is_wms: True se for WMS, False se for API
        range_dados: Tupla com (col_pallet, col_atributo, col_quantity, ultima_linha) para referências de célula
        mapa_cores_pallets: Dicionário {pallet_code: cor_hex} para cores consistentes entre WMS e API
    """
    from openpyxl.utils import get_column_letter
    
    # Cores
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    total_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Usa paleta de cores global para consistência
    paleta_cores = PALETA_CORES_PALLETS
    
    linha_atual = inicio_linha
    
    # Cabeçalho: pallet_code
    col_pallet = get_column_letter(inicio_col)
    ws[f'{col_pallet}{linha_atual}'] = 'pallet_code'
    ws[f'{col_pallet}{linha_atual}'].fill = header_fill
    ws[f'{col_pallet}{linha_atual}'].font = header_font
    ws[f'{col_pallet}{linha_atual}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'{col_pallet}{linha_atual}'].border = thin_border
    
    # Cabeçalhos dos atributos
    for i, atributo in enumerate(atributos):
        col_letra = get_column_letter(inicio_col + 1 + i)
        ws[f'{col_letra}{linha_atual}'] = atributo
        ws[f'{col_letra}{linha_atual}'].fill = header_fill
        ws[f'{col_letra}{linha_atual}'].font = header_font
        ws[f'{col_letra}{linha_atual}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col_letra}{linha_atual}'].border = thin_border
    
    # Cabeçalho ALL
    col_all = get_column_letter(inicio_col + 1 + len(atributos))
    ws[f'{col_all}{linha_atual}'] = 'ALL'
    ws[f'{col_all}{linha_atual}'].fill = header_fill
    ws[f'{col_all}{linha_atual}'].font = header_font
    ws[f'{col_all}{linha_atual}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'{col_all}{linha_atual}'].border = thin_border
    
    linha_atual += 1
    
    # Extrai informações do range (para usar referências A1)
    if range_dados:
        col_pallet_ref, col_atributo_ref, col_quantity_ref, ultima_linha_ref = range_dados
        col_pallet_letra = get_column_letter(col_pallet_ref)
        col_atributo_letra = get_column_letter(col_atributo_ref)
        col_quantity_letra = get_column_letter(col_quantity_ref)
        primeira_linha_dados = 13  # Linha 13 é onde começam os dados (após cabeçalhos)
    
    # Linhas de dados (um pallet por linha)
    for idx_pallet, pallet in enumerate(pallets):
        # Usa mapa de cores se fornecido, senão usa índice na paleta
        if mapa_cores_pallets and pallet in mapa_cores_pallets:
            cor_hex = mapa_cores_pallets[pallet]
        else:
            cor_hex = paleta_cores[idx_pallet % len(paleta_cores)]
        
        cor_pallet = PatternFill(
            start_color=cor_hex,
            end_color=cor_hex,
            fill_type="solid"
        )
        
        # Coluna pallet_code
        ws[f'{col_pallet}{linha_atual}'] = pallet
        ws[f'{col_pallet}{linha_atual}'].fill = cor_pallet
        ws[f'{col_pallet}{linha_atual}'].border = thin_border
        ws[f'{col_pallet}{linha_atual}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Colunas de atributos com fórmulas SUMIFS
        for i, atributo in enumerate(atributos):
            col_letra = get_column_letter(inicio_col + 1 + i)
            
            # Usa SUMPRODUCT com SUBTOTAL para respeitar filtros
            # SUMPRODUCT funciona com filtros quando combinado com verificação de linhas visíveis
            if range_dados:
                # Fórmula que respeita filtros:
                # =SUMPRODUCT((Comparação!$B$13:$B$100="P01_A_01")*(Comparação!$F$13:$F$100="Descartável")*SUBTOTAL(103,OFFSET(Comparação!$E$13,ROW(Comparação!$B$13:$B$100)-ROW(Comparação!$B$13),0,1))*(Comparação!$E$13:$E$100))
                # Versão simplificada usando SUMIFS com ajuste manual para filtros
                # Como SUMIFS não respeita filtros, vamos usar uma abordagem com SUMPRODUCT
                formula = (
                    f'=SUMPRODUCT('
                    f'(Comparação!${col_pallet_letra}${primeira_linha_dados}:${col_pallet_letra}${ultima_linha_ref}="{pallet}")*'
                    f'(Comparação!${col_atributo_letra}${primeira_linha_dados}:${col_atributo_letra}${ultima_linha_ref}="{atributo}")*'
                    f'SUBTOTAL(103,OFFSET(Comparação!${col_quantity_letra}${primeira_linha_dados},ROW(Comparação!${col_quantity_letra}${primeira_linha_dados}:${col_quantity_letra}${ultima_linha_ref})-ROW(Comparação!${col_quantity_letra}${primeira_linha_dados}),0,1))*'
                    f'(Comparação!${col_quantity_letra}${primeira_linha_dados}:${col_quantity_letra}${ultima_linha_ref})'
                    f')'
                )
            else:
                # Fallback para referência estruturada (pode não funcionar no SharePoint)
                formula = f'=SUMIFS({nome_tabela_origem}[quantity],{nome_tabela_origem}[pallet_code],"{pallet}",{nome_tabela_origem}[atributo],"{atributo}")'
            
            ws[f'{col_letra}{linha_atual}'] = formula
            ws[f'{col_letra}{linha_atual}'].fill = cor_pallet
            ws[f'{col_letra}{linha_atual}'].border = thin_border
            ws[f'{col_letra}{linha_atual}'].alignment = Alignment(horizontal='right', vertical='center')
        
        # Coluna ALL (soma de todos os atributos)
        col_letra_inicio = get_column_letter(inicio_col + 1)
        col_letra_fim = get_column_letter(inicio_col + len(atributos))
        ws[f'{col_all}{linha_atual}'] = f'=SUM({col_letra_inicio}{linha_atual}:{col_letra_fim}{linha_atual})'
        ws[f'{col_all}{linha_atual}'].fill = cor_pallet
        ws[f'{col_all}{linha_atual}'].border = thin_border
        ws[f'{col_all}{linha_atual}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'{col_all}{linha_atual}'].font = Font(bold=True)
        
        linha_atual += 1
    
    # Linha ALL (totais)
    ws[f'{col_pallet}{linha_atual}'] = 'ALL'
    ws[f'{col_pallet}{linha_atual}'].fill = total_fill
    ws[f'{col_pallet}{linha_atual}'].font = total_font
    ws[f'{col_pallet}{linha_atual}'].border = thin_border
    ws[f'{col_pallet}{linha_atual}'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Totais por atributo
    for i, atributo in enumerate(atributos):
        col_letra = get_column_letter(inicio_col + 1 + i)
        linha_inicio = inicio_linha + 1
        linha_fim = linha_atual - 1
        ws[f'{col_letra}{linha_atual}'] = f'=SUM({col_letra}{linha_inicio}:{col_letra}{linha_fim})'
        ws[f'{col_letra}{linha_atual}'].fill = total_fill
        ws[f'{col_letra}{linha_atual}'].font = total_font
        ws[f'{col_letra}{linha_atual}'].border = thin_border
        ws[f'{col_letra}{linha_atual}'].alignment = Alignment(horizontal='right', vertical='center')
    
    # Total ALL
    col_letra_inicio = get_column_letter(inicio_col + 1)
    col_letra_fim = get_column_letter(inicio_col + len(atributos))
    ws[f'{col_all}{linha_atual}'] = f'=SUM({col_letra_inicio}{linha_atual}:{col_letra_fim}{linha_atual})'
    ws[f'{col_all}{linha_atual}'].fill = total_fill
    ws[f'{col_all}{linha_atual}'].font = total_font
    ws[f'{col_all}{linha_atual}'].border = thin_border
    ws[f'{col_all}{linha_atual}'].alignment = Alignment(horizontal='right', vertical='center')
    
    return linha_atual


def gerar_cores_por_pallet_tabela_principal(ws) -> dict:
    """
    Gera um mapeamento de pallet_code para cores neutras alternadas para a tabela principal.
    Inclui TODOS os pallets (WMS + API) para garantir cores consistentes.
    
    Args:
        ws: Worksheet da tabela Comparação
        
    Returns:
        Dict com {pallet_code: PatternFill}
    """
    # Usa paleta de cores global para consistência
    paleta_cores = PALETA_CORES_PALLETS
    
    # CRÍTICO: Primeiro coleta todos os pallets do WMS (ordem de aparição)
    # Depois adiciona pallets exclusivos da API
    # Isso garante que o MESMO pallet tenha a MESMA cor em ambas as tabelas
    pallet_codes = []
    
    # Primeiro: todos os pallets do WMS (na ordem de aparição)
    for row in range(13, ws.max_row + 1):
        pallet_code_wms = ws.cell(row=row, column=2).value  # Coluna B = pallet_code WMS
        if pallet_code_wms and pallet_code_wms not in pallet_codes and pallet_code_wms != '':
            pallet_codes.append(pallet_code_wms)
    
    # Segundo: adiciona pallets da API que NÃO existem no WMS
    for row in range(13, ws.max_row + 1):
        pallet_code_api = ws.cell(row=row, column=10).value  # Coluna J = pallet_code API
        if pallet_code_api and pallet_code_api not in pallet_codes and pallet_code_api != '':
            pallet_codes.append(pallet_code_api)
    
    # Mapeia cada pallet_code para uma cor
    cores_mapeadas = {}
    for i, pallet_code in enumerate(pallet_codes):
        cor_hex = paleta_cores[i % len(paleta_cores)]
        cores_mapeadas[pallet_code] = PatternFill(
            start_color=cor_hex,
            end_color=cor_hex,
            fill_type="solid"
        )
    
    return cores_mapeadas


def criar_secao_kpis(ws, header_fill, header_font, thin_border, separador_fill):
    """
    Cria seção de KPIs na parte superior da planilha.
    Usa coluna auxiliar H para verificar matches (mais compatível com SharePoint).
    Responde a filtros usando SUBTOTAL.
    
    Args:
        ws: Worksheet
        header_fill: Preenchimento dos cabeçalhos
        header_font: Fonte dos cabeçalhos
        thin_border: Borda padrão
        separador_fill: Preenchimento do separador
    """
    kpi_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    info_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    kpi_font = Font(bold=True, size=10)
    value_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=11, color="FFFFFF")
    
    # ===========================================
    # LINHA 1: INFORMAÇÕES GERAIS
    # ===========================================
    
    # Mapas Processados
    ws.merge_cells('B1:C1')
    ws['B1'] = 'Mapas Processados:'
    ws['B1'].fill = info_fill
    ws['B1'].font = Font(bold=True, size=10)
    ws['B1'].alignment = Alignment(horizontal='right', vertical='center')
    ws['B1'].border = thin_border
    
    # Fórmula para contar number_mapa únicos (compatível com SharePoint)
    # Usa SUMPRODUCT com comparação de primeira ocorrência
    ws['D1'] = (
        '=SUMPRODUCT(($A$13:$A$5000<>"")/COUNTIF($A$13:$A$5000,$A$13:$A$5000&""))'
    )
    ws['D1'].font = value_font
    ws['D1'].border = thin_border
    ws['D1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['D1'].fill = info_fill
    ws['D1'].number_format = '0'
    
    # Taxa de Acerto Média (percentual médio por number_mapa)
    ws.merge_cells('E1:F1')
    ws['E1'] = 'Taxa Média p/ Mapa:'
    ws['E1'].fill = info_fill
    ws['E1'].font = Font(bold=True, size=10)
    ws['E1'].alignment = Alignment(horizontal='right', vertical='center')
    ws['E1'].border = thin_border
    
    # Calcula percentual médio agrupado por number_mapa
    # Média simples: total de matches / total de linhas com dados
    # Compatível com SharePoint Excel
    ws['G1'] = (
        '=IFERROR(SUMPRODUCT(($A$13:$A$5000<>"")*($H$13:$H$5000=1))/'
        'SUMPRODUCT(($A$13:$A$5000<>"")*1),0)'
    )
    ws['G1'].font = value_font
    ws['G1'].border = thin_border
    ws['G1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['G1'].fill = info_fill
    ws['G1'].number_format = '0.0%'
    
    # ===========================================
    # LINHA 2: TÍTULO DA TABELA KPI
    # ===========================================
    
    ws.merge_cells('B2:D2')
    ws['B2'] = 'KPIs de Comparação'
    ws['B2'].fill = header_fill
    ws['B2'].font = title_font
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'].border = thin_border
    
    # ===========================================
    # LINHA 3: CABEÇALHOS DA TABELA KPI
    # ===========================================
    
    ws['B3'] = 'Nome'
    ws['B3'].fill = header_fill
    ws['B3'].font = title_font
    ws['B3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B3'].border = thin_border
    
    ws['C3'] = 'Valor'
    ws['C3'].fill = header_fill
    ws['C3'].font = title_font
    ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C3'].border = thin_border
    
    ws['D3'] = 'Percentual'
    ws['D3'].fill = header_fill
    ws['D3'].font = title_font
    ws['D3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['D3'].border = thin_border
    
    # ===========================================
    # LINHA 4: DADOS DO KPI
    # ===========================================
    
    # Nome do KPI
    ws['B4'] = 'Linhas Idênticas (WMS = API)'
    ws['B4'].fill = kpi_fill
    ws['B4'].font = kpi_font
    ws['B4'].alignment = Alignment(horizontal='left', vertical='center')
    ws['B4'].border = thin_border
    
    # Valor: Conta linhas onde coluna H = 1 (match perfeito)
    # Usa SUBTOTAL para responder a filtros
    ws['C4'] = (
        '=SUMPRODUCT((SUBTOTAL(103,OFFSET(H13,ROW(H13:H1000)-ROW(H13),0,1)))*(H13:H1000=1))'
    )
    ws['C4'].font = value_font
    ws['C4'].border = thin_border
    ws['C4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C4'].fill = kpi_fill
    
    # Percentual: Linhas idênticas / Total de linhas
    ws['D4'] = (
        '=IFERROR('
        'SUMPRODUCT((SUBTOTAL(103,OFFSET(H13,ROW(H13:H1000)-ROW(H13),0,1)))*(H13:H1000=1))/'
        'SUMPRODUCT((SUBTOTAL(103,OFFSET(A13,ROW(A13:A1000)-ROW(A13),0,1)))*(A13:A1000<>""))'
        ',0)'
    )
    ws['D4'].font = value_font
    ws['D4'].border = thin_border
    ws['D4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['D4'].fill = kpi_fill
    ws['D4'].number_format = '0.0%'
    
    # ===========================================
    # LINHA 5: DESCRIÇÃO
    # ===========================================
    
    ws.merge_cells('B5:D5')
    ws['B5'] = 'Percentual global | Taxa Média = média de acerto por mapa'
    ws['B5'].font = Font(size=9, italic=True, color="666666")
    ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajusta larguras das colunas
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 3  # Coluna auxiliar Match


def criar_tabela_pivot(df_wms: pd.DataFrame, df_api: pd.DataFrame) -> tuple:
    """
    Cria tabelas pivot agrupando por pallet_code e atributo.
    
    Args:
        df_wms: DataFrame com dados WMS
        df_api: DataFrame com dados API
        
    Returns:
        Tupla com (df_pivot_wms, df_pivot_api)
    """
    def criar_pivot(df: pd.DataFrame) -> pd.DataFrame:
        """Cria pivot table com pallet_code nas linhas e atributo nas colunas."""
        # Agrupa por pallet_code e atributo, somando quantities
        pivot = df.groupby(['pallet_code', 'atributo'])['quantity'].sum().unstack(fill_value=0)
        
        # Adiciona coluna ALL (total por pallet)
        pivot['ALL'] = pivot.sum(axis=1)
        
        # Adiciona linha ALL (total por atributo)
        pivot.loc['ALL'] = pivot.sum(axis=0)
        
        # Ordena colunas alfabeticamente, com ALL no final
        colunas = sorted([col for col in pivot.columns if col != 'ALL']) + ['ALL']
        pivot = pivot[colunas]
        
        # Reset index para ter pallet_code como coluna
        pivot = pivot.reset_index()
        
        return pivot
    
    df_pivot_wms = criar_pivot(df_wms)
    df_pivot_api = criar_pivot(df_api)
    
    return df_pivot_wms, df_pivot_api


def criar_aba_pivot(writer, df_pivot_wms: pd.DataFrame, df_pivot_api: pd.DataFrame):
    """
    Cria aba com tabelas pivot lado a lado.
    
    Args:
        writer: ExcelWriter object
        df_pivot_wms: Pivot table WMS
        df_pivot_api: Pivot table API
    """
    # Cria separador
    max_rows = max(len(df_pivot_wms), len(df_pivot_api))
    df_separador = pd.DataFrame({'': [''] * max_rows})
    
    # Concatena horizontalmente: WMS + separador + API
    df_pivot_final = pd.concat([df_pivot_wms, df_separador, df_pivot_api], axis=1)
    df_pivot_final = df_pivot_final.fillna('')
    
    # Salva na aba 'Por Tipo'
    df_pivot_final.to_excel(writer, sheet_name='Por Tipo', index=False, startrow=2)


def formatar_aba_pivot(filepath: Path):
    """
    Aplica formatação à aba Por Tipo.
    
    Args:
        filepath: Caminho do arquivo Excel
    """
    wb = load_workbook(filepath)
    
    if 'Por Tipo' not in wb.sheetnames:
        return
    
    ws = wb['Por Tipo']
    
    # Cores
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    separador_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    total_font = Font(bold=True, size=11)
    
    # Borda
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    num_colunas = ws.max_column
    
    # Encontra coluna separadora (coluna vazia '')
    separador_col = None
    for col in range(1, num_colunas + 1):
        if ws.cell(row=3, column=col).value == '':
            separador_col = col
            break
    
    # Linha 2: Cabeçalhos principais (WMS e RESULTADO)
    if separador_col:
        ws.merge_cells(f'A2:{chr(64 + separador_col - 1)}2')
        ws['A2'] = 'WMS'
        ws['A2'].fill = header_fill
        ws['A2'].font = header_font
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.cell(row=2, column=separador_col).fill = separador_fill
        
        ws.merge_cells(f'{chr(64 + separador_col + 1)}2:{chr(64 + num_colunas)}2')
        ws.cell(row=2, column=separador_col + 1).value = 'RESULTADO'
        ws.cell(row=2, column=separador_col + 1).fill = header_fill
        ws.cell(row=2, column=separador_col + 1).font = header_font
        ws.cell(row=2, column=separador_col + 1).alignment = Alignment(horizontal='center', vertical='center')
    
    # Linha 3: Cabeçalhos das colunas
    for col in range(1, num_colunas + 1):
        cell = ws.cell(row=3, column=col)
        if col != separador_col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.fill = separador_fill
    
    # Formata coluna separadora
    if separador_col:
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=separador_col).fill = separador_fill
    
    # Aplica bordas e formata células de dados
    for row in range(4, ws.max_row + 1):
        # Verifica se é linha ALL (última linha)
        is_total_row = (ws.cell(row=row, column=1).value == 'ALL')
        
        for col in range(1, num_colunas + 1):
            cell = ws.cell(row=row, column=col)
            
            if col != separador_col:
                cell.border = thin_border
                
                # Formata linha ALL
                if is_total_row:
                    cell.fill = total_fill
                    cell.font = total_font
                
                # Alinha números à direita (colunas de quantidade)
                if col > 1 or (separador_col and col > separador_col + 1):
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')
    
    # Destaca coluna ALL (última coluna de cada tabela)
    if separador_col:
        # ALL do WMS
        all_col_wms = separador_col - 1
        # ALL do API
        all_col_api = num_colunas
        
        for row in range(4, ws.max_row + 1):
            ws.cell(row=row, column=all_col_wms).font = Font(bold=True)
            ws.cell(row=row, column=all_col_api).font = Font(bold=True)
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 18  # pallet_code
    if separador_col:
        ws.column_dimensions[chr(64 + separador_col)].width = 2  # separador
        ws.column_dimensions[chr(64 + separador_col + 1)].width = 18  # pallet_code API
        
        # Outras colunas (atributos)
        for col in range(2, num_colunas + 1):
            if col != separador_col and col != separador_col + 1:
                ws.column_dimensions[chr(64 + col)].width = 12
    else:
        # Se não houver separador, formata todas as colunas
        for col in range(2, num_colunas + 1):
            ws.column_dimensions[chr(64 + col)].width = 12
    
    # Congela painéis
    ws.freeze_panes = 'A4'
    
    wb.save(filepath)


def formatar_excel(filepath: Path, atributos_wms=None, atributos_api=None, pallets_wms=None, pallets_api=None, is_append=False, wms_file=None, api_file=None, criar_aba_nao_paletizados_flag=True):
    """
    Aplica formatação ao Excel: cabeçalhos, bordas, alinhamento, cores.
    Cria tabela dinâmica Por Tipo com fórmulas.
    
    Args:
        filepath: Caminho do arquivo Excel
        atributos_wms: Lista de atributos únicos WMS
        atributos_api: Lista de atributos únicos API
        pallets_wms: Lista de pallets únicos WMS
        pallets_api: Lista de pallets únicos API
        is_append: Se True, apenas atualiza formatação sem recriar estruturas
        wms_file: Caminho do arquivo WMS TXT (para extrair não paletizados)
        api_file: Caminho do arquivo API TXT (para extrair não paletizados)
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo
    
    wb = load_workbook(filepath)
    ws = wb['Comparação']
    
    # Cores
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    separador_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Borda
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ====================
    # KPIs (Linhas 2-9)
    # ====================
    
    # Sempre cria/atualiza KPIs
    criar_secao_kpis(ws, header_fill, header_font, thin_border, separador_fill)
    
    # ====================
    # TABELAS (A partir da linha 11)
    # ====================
    
    num_colunas = ws.max_column
    
    # Linha 11: Cabeçalhos principais (WMS e RESULTADO)
    ws.merge_cells('A11:G11')
    ws['A11'] = 'WMS'
    ws['A11'].fill = header_fill
    ws['A11'].font = header_font
    ws['A11'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Coluna separadora
    ws['H11'].fill = separador_fill
    
    ws.merge_cells('I11:O11')
    ws['I11'] = 'RESULTADO'
    ws['I11'].fill = header_fill
    ws['I11'].font = header_font
    ws['I11'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Linha 12: Cabeçalhos das colunas (já existente do DataFrame)
    for col in range(1, num_colunas + 1):
        cell = ws.cell(row=12, column=col)
        if col != 8:  # Não formata a coluna separadora
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.fill = separador_fill
    
    # Formata coluna separadora (H) - ANTES de adicionar fórmulas
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=8)
        cell.fill = separador_fill
        # Não sobrescreve as fórmulas que serão adicionadas depois
    
    # Adiciona cabeçalho na coluna H (linha 12)
    ws['H12'] = 'Match'
    ws['H12'].fill = separador_fill
    ws['H12'].font = Font(size=8, color="666666")
    ws['H12'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Adiciona fórmula de match na coluna H (linhas 13+)
    # Verifica se existe linha no API com mesmos valores de number_mapa, pallet_code, product_code e quantity
    # Usa TRIM para remover espaços extras que possam atrapalhar a comparação
    print(f"\n🔍 Aplicando fórmulas de match para {ws.max_row - 12} linhas...")
    formulas_aplicadas = 0
    for row in range(13, ws.max_row + 1):
        # Mantém o fill cinza e aplica a fórmula
        cell = ws[f'H{row}']
        cell.value = (
            f'=IF(TRIM(A{row})<>"",IF(COUNTIFS($I$13:$I$5000,TRIM(A{row}),$J$13:$J$5000,TRIM(B{row}),'
            f'$K$13:$K$5000,TRIM(C{row}),$M$13:$M$5000,E{row})>0,1,""),"")'
        )
        cell.fill = separador_fill
        cell.font = Font(size=8)
        cell.alignment = Alignment(horizontal='center')
        formulas_aplicadas += 1
        
        # Debug: Mostra primeiras e últimas 3 fórmulas
        if row <= 15 or row >= ws.max_row - 2:
            pallet_wms = ws.cell(row=row, column=2).value
            print(f"   H{row}: Pallet={pallet_wms}, Fórmula aplicada")
    
    print(f"✅ {formulas_aplicadas} fórmulas aplicadas de H13 a H{ws.max_row}")
    
    # Gera mapeamento de cores por pallet_code
    cores_por_pallet = gerar_cores_por_pallet_tabela_principal(ws)
    
    # Aplica bordas, alinhamento e cores nas células de dados (a partir da linha 13)
    for row in range(13, ws.max_row + 1):
        # Pega o pallet_code da linha WMS (coluna B) e API (coluna J)
        pallet_code_wms = ws.cell(row=row, column=2).value
        pallet_code_api = ws.cell(row=row, column=10).value  # Coluna J = 10
        
        # Pega as cores de cada pallet
        cor_pallet_wms = cores_por_pallet.get(pallet_code_wms)
        cor_pallet_api = cores_por_pallet.get(pallet_code_api)
        
        for col in range(1, num_colunas + 1):
            cell = ws.cell(row=row, column=col)
            if col != 8:  # Não formata a coluna separadora
                cell.border = thin_border
                
                # Aplica cor baseada no pallet_code de cada lado
                if col <= 7:  # Colunas WMS (A-G)
                    if cor_pallet_wms:
                        cell.fill = cor_pallet_wms
                elif col >= 9:  # Colunas API (I-O)
                    if cor_pallet_api:
                        cell.fill = cor_pallet_api
                
                # Alinha números à direita
                if col in [1, 5, 7, 9, 13, 15]:  # Colunas de number_mapa, quantity e ocupacao
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 12  # number_mapa WMS
    ws.column_dimensions['B'].width = 18  # pallet_code WMS
    ws.column_dimensions['C'].width = 14  # product_code WMS
    ws.column_dimensions['D'].width = 50  # product_name WMS
    ws.column_dimensions['E'].width = 10  # quantity WMS
    ws.column_dimensions['F'].width = 14  # atributo WMS
    ws.column_dimensions['G'].width = 10  # ocupacao WMS
    ws.column_dimensions['H'].width = 2   # separador
    ws.column_dimensions['I'].width = 12  # number_mapa API
    ws.column_dimensions['J'].width = 18  # pallet_code API
    ws.column_dimensions['K'].width = 14  # product_code API
    ws.column_dimensions['L'].width = 50  # product_name API
    ws.column_dimensions['M'].width = 10  # quantity API
    ws.column_dimensions['N'].width = 14  # atributo API
    ws.column_dimensions['O'].width = 10  # ocupacao API
    
    # Adiciona AutoFilter nos cabeçalhos da tabela (linha 12)
    ws.auto_filter.ref = f"A12:O{ws.max_row}"
    
    # Remove tabelas antigas se existirem (para evitar conflitos)
    if not is_append:
        # Converte em Tabela do Excel para permitir referências dinâmicas
        # Tabela WMS (colunas A-G)
        if ws.max_row > 12:
            tab_wms = Table(displayName="TabelaWMS", ref=f"A12:G{ws.max_row}")
            style_wms = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab_wms.tableStyleInfo = style_wms
            ws.add_table(tab_wms)
            
            # Tabela API (colunas I-O)
            tab_api = Table(displayName="TabelaAPI", ref=f"I12:O{ws.max_row}")
            style_api = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab_api.tableStyleInfo = style_api
            ws.add_table(tab_api)
    else:
        # Se é append, atualiza os ranges das tabelas existentes
        print(f"  Atualizando ranges das tabelas para linha {ws.max_row}")
        for table in ws.tables.values():
            if table.name == "TabelaWMS":
                table.ref = f"A12:G{ws.max_row}"
                print(f"    TabelaWMS: {table.ref}")
            elif table.name == "TabelaAPI":
                table.ref = f"I12:O{ws.max_row}"
                print(f"    TabelaAPI: {table.ref}")
    
    # Congela painéis (congela até linha 12 - cabeçalhos das colunas)
    ws.freeze_panes = 'A13'
    
    # Cria aba "Por Tipo" com tabelas dinâmicas usando fórmulas
    if atributos_wms and pallets_wms and atributos_api and pallets_api:
        # Remove aba anterior se existir
        if "Por Tipo" in wb.sheetnames:
            del wb["Por Tipo"]
        
        ws_pivot = wb.create_sheet("Por Tipo")
        
        # Calcula última linha da tabela de dados
        ultima_linha_dados = ws.max_row
        
        # CRÍTICO: USA O MESMO MAPA DE CORES da tabela principal (cores_por_pallet)
        # Isso garante que o MESMO pallet tenha a MESMA cor em WMS e API
        # Converte PatternFill -> string hex para passar para criar_tabela_dinamica_com_formulas
        mapa_cores_pallets = {}
        for pallet_code, pattern_fill in cores_por_pallet.items():
            # Extrai cor hex do PatternFill
            cor_hex = pattern_fill.start_color.rgb if hasattr(pattern_fill.start_color, 'rgb') else pattern_fill.start_color
            # Remove o prefixo 'FF' se existir (alpha channel do ARGB)
            if isinstance(cor_hex, str) and len(cor_hex) == 8 and cor_hex.startswith('FF'):
                cor_hex = cor_hex[2:]
            mapa_cores_pallets[pallet_code] = cor_hex
        
        print(f"  🎨 Usando mapa de cores UNIFICADO na aba Por Tipo ({len(mapa_cores_pallets)} pallets)")
        
        # CRÍTICO: Calcula posições dinâmicas baseado no número de atributos
        # WMS: 1 coluna pallet + N colunas atributos + 1 coluna ALL = N+2 colunas
        num_colunas_wms = 1 + len(atributos_wms) + 1  # pallet_code + atributos + ALL
        col_separador = num_colunas_wms + 1  # Próxima coluna após WMS
        inicio_col_api = col_separador + 1  # Próxima coluna após separador
        
        # Calcula letras das colunas para merge cells
        from openpyxl.utils import get_column_letter
        ultima_col_wms = get_column_letter(num_colunas_wms)
        col_separador_letra = get_column_letter(col_separador)
        primeira_col_api = get_column_letter(inicio_col_api)
        num_colunas_api = 1 + len(atributos_api) + 1
        ultima_col_api = get_column_letter(inicio_col_api + num_colunas_api - 1)
        
        print(f"  📊 Layout dinâmico da aba Por Tipo:")
        print(f"     WMS: A-{ultima_col_wms} ({num_colunas_wms} colunas)")
        print(f"     Separador: {col_separador_letra}")
        print(f"     API: {primeira_col_api}-{ultima_col_api} ({num_colunas_api} colunas)")
        
        # Título WMS
        ws_pivot.merge_cells(f'A2:{ultima_col_wms}2')
        ws_pivot['A2'] = 'WMS'
        ws_pivot['A2'].fill = header_fill
        ws_pivot['A2'].font = header_font
        ws_pivot['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Cria tabela WMS com fórmulas (começa na linha 3, coluna 1)
        # Range WMS: B=pallet_code(2), F=atributo(6), E=quantity(5)
        range_wms = (2, 6, 5, ultima_linha_dados)  # (col_pallet, col_atributo, col_quantity, ultima_linha)
        ultima_linha_wms = criar_tabela_dinamica_com_formulas(
            ws_pivot, 'TabelaWMS', 3, 1, pallets_wms, atributos_wms, is_wms=True, range_dados=range_wms, mapa_cores_pallets=mapa_cores_pallets
        )
        
        # Coluna separadora (dinâmica)
        for row in range(2, ultima_linha_wms + 1):
            ws_pivot.cell(row=row, column=col_separador).fill = separador_fill
        ws_pivot.column_dimensions[col_separador_letra].width = 2
        
        # Título API
        ws_pivot.merge_cells(f'{primeira_col_api}2:{ultima_col_api}2')
        ws_pivot.cell(row=2, column=inicio_col_api).value = 'RESULTADO'
        ws_pivot.cell(row=2, column=inicio_col_api).fill = header_fill
        ws_pivot.cell(row=2, column=inicio_col_api).font = header_font
        ws_pivot.cell(row=2, column=inicio_col_api).alignment = Alignment(horizontal='center', vertical='center')
        
        # Cria tabela API com fórmulas (começa na linha 3, coluna dinâmica)
        # Range API: J=pallet_code(10), N=atributo(14), M=quantity(13)
        range_api = (10, 14, 13, ultima_linha_dados)  # (col_pallet, col_atributo, col_quantity, ultima_linha)
        criar_tabela_dinamica_com_formulas(
            ws_pivot, 'TabelaAPI', 3, inicio_col_api, pallets_api, atributos_api, is_wms=False, range_dados=range_api, mapa_cores_pallets=mapa_cores_pallets
        )
        
        # Ajusta larguras das colunas dinamicamente
        ws_pivot.column_dimensions['A'].width = 18  # pallet_code WMS
        ws_pivot.column_dimensions[primeira_col_api].width = 18  # pallet_code API
        
        # Largura das colunas de atributos WMS (B em diante)
        for i in range(len(atributos_wms) + 1):  # Atributos + ALL
            col_letra = get_column_letter(2 + i)  # Começa da coluna B (2)
            ws_pivot.column_dimensions[col_letra].width = 12
        
        # Largura das colunas de atributos API (após o separador)
        for i in range(len(atributos_api) + 1):  # Atributos + ALL
            col_letra = get_column_letter(inicio_col_api + 1 + i)  # Pula pallet_code
            ws_pivot.column_dimensions[col_letra].width = 12
        
        # Freeze panes
        ws_pivot.freeze_panes = 'A4'
    
    # Cria aba "Consolidado" com resumo por mapa
    criar_aba_consolidado(wb, ws)
    
    # Cria aba "Não Paletizados" APENAS se criar_aba_nao_paletizados_flag=True
    # Em modo batch, só cria no último mapa para agregar TODOS os produtos
    if criar_aba_nao_paletizados_flag:
        print(f"\n[DEBUG] Criando aba Nao Paletizados com todos os mapas processados")
        criar_aba_nao_paletizados(wb, ws)
        print(f"[DEBUG] criar_aba_nao_paletizados retornou")
    else:
        print(f"\n[DEBUG] Pulando criacao da aba Nao Paletizados (sera criada no final do batch)")
    
    print(f"\n[DEBUG] Salvando workbook em: {filepath}")
    print(f"[DEBUG] Abas no workbook: {wb.sheetnames}")
    wb.save(filepath)
    print(f"[DEBUG] Workbook salvo!")


def criar_aba_consolidado_binpack(wb, ws_comparacao):
    """
    Cria aba 'Consolidado' com resumo por number_mapa.
    """
    from openpyxl.utils import get_column_letter

    # Remove aba anterior se existir
    if "Consolidado" in wb.sheetnames:
        del wb["Consolidado"]

    ws = wb.create_sheet("Consolidado", 0)  # Cria como primeira aba

    # Cores
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:J1')
    ws['A1'] = 'Resumo Consolidado por Mapa'
    ws['A1'].fill = header_fill
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].border = thin_border

    # Cabeçalhos (linha 3) - adicionada coluna 'Binpack' no final
    headers = [
        'number_mapa',
        'Similaridade (%)',
        'Ocupação Total WMS',
        'Ocupação Total API',
        'Total Itens WMS',
        'Total Itens API',
        'Perc. Não Paletizados WMS',
        'Perc. Não Paletizados API',
        'Tempo Execução (segs)',
        'Binpack'
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Extrai number_mapas únicos da aba Comparação
    mapas_unicos = []
    linha = 13
    while True:
        valor = ws_comparacao.cell(row=linha, column=1).value
        if not valor or valor == '':
            break
        if valor not in mapas_unicos:
            mapas_unicos.append(valor)
        linha += 1

    print(f"\n  📊 Criando aba Consolidado com {len(mapas_unicos)} mapas únicos")
    print(f"  Mapas únicos encontrados: {mapas_unicos}")

    # (mantém lógica existente para buscar durations_map...)
    # ... (aqui entra o código já presente que popula durations_map) ...
    durations_map = {}
    try:
        # tentativa de leitura do batch_summary_*.json ou fallback log (mesma lógica já implementada)
        global _MAPEAMENTO_ARQUIVOS
        candidate_root = None
        if _MAPEAMENTO_ARQUIVOS:
            any_entry = next(iter(_MAPEAMENTO_ARQUIVOS.values()))
            wms_path = Path(any_entry[0])
            candidate_root = wms_path.parent.parent
        else:
            candidate_root = Path(__file__).parent.parent

        if candidate_root and candidate_root.exists():
            js_files = sorted(candidate_root.glob('batch_summary_*.json'), key=lambda p: p.stat().st_mtime, reverse=True)
            if js_files:
                try:
                    with open(js_files[0], 'r', encoding='utf-8') as f:
                        summary = json.load(f)
                    for r in summary.get('results', []):
                        map_name = str(r.get('map_name') or r.get('mapName') or r.get('map_number') or '')
                        s = r.get('start_time') or r.get('startTime') or None
                        e = r.get('end_time') or r.get('endTime') or None
                        dur = None
                        try:
                            if s and e:
                                from datetime import datetime
                                dt_s = datetime.fromisoformat(s)
                                dt_e = datetime.fromisoformat(e)
                                dur = (dt_e - dt_s).total_seconds()
                        except Exception:
                            dur = None
                        if dur is None:
                            dur = r.get('duration_seconds') or r.get('durationSeconds') or None
                        if dur is not None and map_name:
                            try:
                                durations_map[map_name] = float(dur)
                            except Exception:
                                durations_map[map_name] = dur
                    print(f"  [INFO] Usando {js_files[0].name} para preencher tempos de execução")
                except Exception as ex:
                    print(f"  [WARN] Erro ao ler {js_files[0].name}: {ex}")
            else:
                log_files = sorted(candidate_root.glob('batch_processing_*.log'), key=lambda p: p.stat().st_mtime, reverse=True)
                if log_files:
                    log_text = log_files[0].read_text(encoding='utf-8', errors='ignore')
                    import re
                    cur_map = None
                    for line in log_text.splitlines():
                        if 'PROCESSANDO MAPA' in line:
                            m = re.search(r'PROCESSANDO MAPA[:\s]*([^\s:]+)', line)
                            if m:
                                cur_map = m.group(1).strip()
                        m2 = re.search(r'Tempo de processamento:\s*([\d\.]+)s', line)
                        if m2 and cur_map:
                            try:
                                durations_map[cur_map] = float(m2.group(1))
                            except Exception:
                                pass
                            cur_map = None
                    print(f"  [INFO] Usando {log_files[0].name} (fallback) para preencher tempos de execução")
    except Exception as e:
        print(f"  [WARN] Não foi possível localizar/parsear arquivos de batch: {e}")

    # Calcula similaridade para cada mapa para poder ordenar
    mapas_com_similaridade = []
    for mapa in mapas_unicos:
        matches = 0
        total = 0
        linha = 13
        while True:
            valor_mapa = ws_comparacao.cell(row=linha, column=1).value
            if not valor_mapa or valor_mapa == '':
                break
            if str(valor_mapa) == str(mapa):
                total += 1
                match_val = ws_comparacao.cell(row=linha, column=8).value
                if match_val == 1:
                    matches += 1
            linha += 1

        similaridade = (matches / total) if total > 0 else 0
        mapas_com_similaridade.append((mapa, similaridade))

    mapas_com_similaridade.sort(key=lambda x: x[1], reverse=True)
    print(f"  Mapas ordenados por similaridade: {[(m, f'{s*100:.1f}%') for m, s in mapas_com_similaridade]}")

    # Adiciona dados (uma linha por mapa, já ordenado)
    linha_atual = 4
    for mapa, _ in mapas_com_similaridade:
        # Coluna A: number_mapa
        ws.cell(row=linha_atual, column=1).value = mapa
        ws.cell(row=linha_atual, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=1).border = thin_border

        # Coluna B: Similaridade (%)
        ws.cell(row=linha_atual, column=2).value = (
            f"=IFERROR(SUMIFS('Comparação'!$H$13:$H$5000,'Comparação'!$A$13:$A$5000,A{linha_atual})/"
            f"COUNTIF('Comparação'!$A$13:$A$5000,A{linha_atual}),0)"
        )
        ws.cell(row=linha_atual, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=2).border = thin_border
        ws.cell(row=linha_atual, column=2).number_format = '0.0%'

        # Coluna C: Ocupação Total WMS
        ws.cell(row=linha_atual, column=3).value = (
            f"=SUMIF('Comparação'!$A$13:$A$5000,A{linha_atual},'Comparação'!$G$13:$G$5000)"
        )
        ws.cell(row=linha_atual, column=3).alignment = Alignment(horizontal='right')
        ws.cell(row=linha_atual, column=3).border = thin_border
        ws.cell(row=linha_atual, column=3).number_format = '0.0'

        # Coluna D: Ocupação Total API
        ws.cell(row=linha_atual, column=4).value = (
            f"=SUMIF('Comparação'!$I$13:$I$5000,A{linha_atual},'Comparação'!$O$13:$O$5000)"
        )
        ws.cell(row=linha_atual, column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=linha_atual, column=4).border = thin_border
        ws.cell(row=linha_atual, column=4).number_format = '0.0'

        # Coluna E: Total Itens WMS
        ws.cell(row=linha_atual, column=5).value = (
            f"=SUMIF('Comparação'!$A$13:$A$5000,A{linha_atual},'Comparação'!$E$13:$E$5000)"
        )
        ws.cell(row=linha_atual, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=5).border = thin_border
        ws.cell(row=linha_atual, column=5).number_format = '0'

        # Coluna F: Total Itens API
        ws.cell(row=linha_atual, column=6).value = (
            f"=SUMIF('Comparação'!$I$13:$I$5000,A{linha_atual},'Comparação'!$M$13:$M$5000)"
        )
        ws.cell(row=linha_atual, column=6).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=6).border = thin_border
        ws.cell(row=linha_atual, column=6).number_format = '0'

        # Coluna G: Perc. Não Paletizados WMS
        ws.cell(row=linha_atual, column=7).value = (
            f"=IFERROR(COUNTIF('Não Paletizados'!$A$5:$A$5000,A{linha_atual})/E{linha_atual},0)"
        )
        ws.cell(row=linha_atual, column=7).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=7).border = thin_border
        ws.cell(row=linha_atual, column=7).number_format = '0.0%'

        # Coluna H: Perc. Não Paletizados API
        ws.cell(row=linha_atual, column=8).value = (
            f"=IFERROR(COUNTIF('Não Paletizados'!$H$5:$H$5000,A{linha_atual})/F{linha_atual},0)"
        )
        ws.cell(row=linha_atual, column=8).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=8).border = thin_border
        ws.cell(row=linha_atual, column=8).number_format = '0.0%'

        # Coluna I: Tempo Execução (segs) - valor direto se disponível no summary/log
        tempo = durations_map.get(str(mapa), None)
        if tempo is not None:
            ws.cell(row=linha_atual, column=9).value = float(tempo)
            ws.cell(row=linha_atual, column=9).number_format = '0.00'
        else:
            ws.cell(row=linha_atual, column=9).value = ''
        ws.cell(row=linha_atual, column=9).alignment = Alignment(horizontal='right')
        ws.cell(row=linha_atual, column=9).border = thin_border

        # Coluna J: Binpack - verifica na aba Comparação se existe atributo contendo 'market' (WMS side coluna F)
        is_binpack = False
        scan_row = 13
        while True:
            val = ws_comparacao.cell(row=scan_row, column=1).value
            if not val or val == '':
                break
            if str(val) == str(mapa):
                atributo_val = ws_comparacao.cell(row=scan_row, column=6).value or ''
                if 'market' in str(atributo_val).lower():
                    is_binpack = True
                    break
            scan_row += 1
        ws.cell(row=linha_atual, column=10).value = 'Sim' if is_binpack else 'Não'
        ws.cell(row=linha_atual, column=10).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=10).border = thin_border

        linha_atual += 1

    print(f"  ✅ {len(mapas_com_similaridade)} linhas de dados criadas (linhas 4 a {linha_atual-1})")

    # Cria Table Excel para permitir filtros e ordenação
    from openpyxl.worksheet.table import Table, TableStyleInfo

    # Define range da tabela (sem a linha de totais)
    tab_range = f"A3:J{linha_atual-1}"
    tab = Table(displayName="TabelaConsolidado", ref=tab_range)

    # Estilo da tabela
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style

    ws.add_table(tab)
    print(f"  ✅ Table Excel criada: {tab_range}")

    # Linha de totais (fora da tabela)
    ws.cell(row=linha_atual, column=1).value = 'TOTAL'
    ws.cell(row=linha_atual, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=1).font = Font(bold=True)
    ws.cell(row=linha_atual, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=1).border = thin_border

    # Similaridade TOTAL
    ws.cell(row=linha_atual, column=2).value = (
        f"=IFERROR(SUMPRODUCT(('Comparação'!$A$13:$A$5000<>\"\")*('Comparação'!$H$13:$H$5000=1))/"
        f"SUMPRODUCT(('Comparação'!$A$13:$A$5000<>\"\")* 1),0)"
    )
    ws.cell(row=linha_atual, column=2).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=2).font = Font(bold=True)
    ws.cell(row=linha_atual, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=2).border = thin_border
    ws.cell(row=linha_atual, column=2).number_format = '0.0%'

    # Totais / médias existentes (mantidos)
    ws.cell(row=linha_atual, column=3).value = f'=SUM(C4:C{linha_atual-1})'
    ws.cell(row=linha_atual, column=3).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=3).font = Font(bold=True)
    ws.cell(row=linha_atual, column=3).alignment = Alignment(horizontal='right')
    ws.cell(row=linha_atual, column=3).border = thin_border
    ws.cell(row=linha_atual, column=3).number_format = '0.0'

    ws.cell(row=linha_atual, column=4).value = f'=SUM(D4:D{linha_atual-1})'
    ws.cell(row=linha_atual, column=4).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=4).font = Font(bold=True)
    ws.cell(row=linha_atual, column=4).alignment = Alignment(horizontal='right')
    ws.cell(row=linha_atual, column=4).border = thin_border
    ws.cell(row=linha_atual, column=4).number_format = '0.0'

    ws.cell(row=linha_atual, column=5).value = f'=SUM(E4:E{linha_atual-1})'
    ws.cell(row=linha_atual, column=5).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=5).font = Font(bold=True)
    ws.cell(row=linha_atual, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=5).border = thin_border
    ws.cell(row=linha_atual, column=5).number_format = '0'

    ws.cell(row=linha_atual, column=6).value = f'=SUM(F4:F{linha_atual-1})'
    ws.cell(row=linha_atual, column=6).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=6).font = Font(bold=True)
    ws.cell(row=linha_atual, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=6).border = thin_border
    ws.cell(row=linha_atual, column=6).number_format = '0'

    ws.cell(row=linha_atual, column=7).value = f'=AVERAGE(G4:G{linha_atual-1})'
    ws.cell(row=linha_atual, column=7).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=7).font = Font(bold=True)
    ws.cell(row=linha_atual, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=7).border = thin_border
    ws.cell(row=linha_atual, column=7).number_format = '0.0%'

    ws.cell(row=linha_atual, column=8).value = f'=AVERAGE(H4:H{linha_atual-1})'
    ws.cell(row=linha_atual, column=8).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=8).font = Font(bold=True)
    ws.cell(row=linha_atual, column=8).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=8).border = thin_border
    ws.cell(row=linha_atual, column=8).number_format = '0.0%'

    # Tempo (coluna I) média
    ws.cell(row=linha_atual, column=9).value = f"=IFERROR(AVERAGE(I4:I{linha_atual-1}),0)"
    ws.cell(row=linha_atual, column=9).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=9).font = Font(bold=True)
    ws.cell(row=linha_atual, column=9).alignment = Alignment(horizontal='right')
    ws.cell(row=linha_atual, column=9).border = thin_border
    ws.cell(row=linha_atual, column=9).number_format = '0.00'

    # Binpack coluna (coluna J) - total de mapas binpack (opcional): conta 'Sim'
    ws.cell(row=linha_atual, column=10).value = f' =COUNTIF(J4:J{linha_atual-1},"Sim")'
    ws.cell(row=linha_atual, column=10).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=10).font = Font(bold=True)
    ws.cell(row=linha_atual, column=10).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=10).border = thin_border

    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 12

    # Freeze panes
    ws.freeze_panes = 'A4'

    print(f"  ✅ Aba Consolidado criada com {len(mapas_unicos)} mapas")

def criar_aba_consolidado(wb, ws_comparacao):
    """
    Cria aba 'Consolidado' com resumo por number_mapa.
    """
    from openpyxl.utils import get_column_letter
    import json

    # Remove aba anterior se existir
    if "Consolidado" in wb.sheetnames:
        del wb["Consolidado"]

    ws = wb.create_sheet("Consolidado", 0)  # Cria como primeira aba

    # Cores
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:I1')
    ws['A1'] = 'Resumo Consolidado por Mapa'
    ws['A1'].fill = header_fill
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].border = thin_border

    # Cabeçalhos (linha 3)
    headers = [
        'number_mapa',
        'Similaridade (%)',
        'Ocupação Total WMS',
        'Ocupação Total API',
        'Total Itens WMS',
        'Total Itens API',
        'Perc. Não Paletizados WMS',
        'Perc. Não Paletizados API',
        'Tempo Execução (segs)'
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Extrai number_mapas únicos da aba Comparação
    mapas_unicos = []
    linha = 13
    while True:
        valor = ws_comparacao.cell(row=linha, column=1).value
        if not valor or valor == '':
            break
        if valor not in mapas_unicos:
            mapas_unicos.append(valor)
        linha += 1

    print(f"\n  📊 Criando aba Consolidado com {len(mapas_unicos)} mapas únicos")
    print(f"  Mapas únicos encontrados: {mapas_unicos}")

    # Tenta localizar arquivo summary/log do batch para extrair durações
    durations_map = {}
    try:
        global _MAPEAMENTO_ARQUIVOS
        candidate_root = None
        if _MAPEAMENTO_ARQUIVOS:
            any_entry = next(iter(_MAPEAMENTO_ARQUIVOS.values()))
            wms_path = Path(any_entry[0])
            candidate_root = wms_path.parent.parent
        else:
            candidate_root = Path(__file__).parent.parent

        if candidate_root and candidate_root.exists():
            # procura batch_summary_*.json (preferencial)
            js_files = sorted(candidate_root.glob('batch_summary_*.json'), key=lambda p: p.stat().st_mtime, reverse=True)
            if js_files:
                try:
                    with open(js_files[0], 'r', encoding='utf-8') as f:
                        summary = json.load(f)
                    for r in summary.get('results', []):
                        map_name = str(r.get('map_name') or r.get('mapName') or r.get('map_number') or '')
                        # prefer compute from start/end if available
                        s = r.get('start_time') or r.get('startTime') or None
                        e = r.get('end_time') or r.get('endTime') or None
                        dur = None
                        try:
                            if s and e:
                                from datetime import datetime
                                dt_s = datetime.fromisoformat(s)
                                dt_e = datetime.fromisoformat(e)
                                dur = (dt_e - dt_s).total_seconds()
                        except Exception:
                            dur = None
                        # fallback to duration_seconds field if present
                        if dur is None:
                            dur = r.get('duration_seconds') or r.get('durationSeconds') or None
                        if dur is not None and map_name:
                            try:
                                durations_map[map_name] = float(dur)
                            except Exception:
                                durations_map[map_name] = dur
                    print(f"  [INFO] Usando {js_files[0].name} para preencher tempos de execução")
                except Exception as ex:
                    print(f"  [WARN] Erro ao ler {js_files[0].name}: {ex}")
            else:
                # fallback para log parsing (mais frágil)
                log_files = sorted(candidate_root.glob('batch_processing_*.log'), key=lambda p: p.stat().st_mtime, reverse=True)
                if log_files:
                    log_text = log_files[0].read_text(encoding='utf-8', errors='ignore')
                    import re
                    cur_map = None
                    for line in log_text.splitlines():
                        if 'PROCESSANDO MAPA' in line:
                            m = re.search(r'PROCESSANDO MAPA[:\s]*([^\s:]+)', line)
                            if m:
                                cur_map = m.group(1).strip()
                        m2 = re.search(r'Tempo de processamento:\s*([\d\.]+)s', line)
                        if m2 and cur_map:
                            try:
                                durations_map[cur_map] = float(m2.group(1))
                            except Exception:
                                pass
                            cur_map = None
                    print(f"  [INFO] Usando {log_files[0].name} (fallback) para preencher tempos de execução")
    except Exception as e:
        print(f"  [WARN] Não foi possível localizar/parsear arquivos de batch: {e}")

    # Calcula similaridade para cada mapa para poder ordenar
    mapas_com_similaridade = []
    for mapa in mapas_unicos:
        matches = 0
        total = 0
        linha = 13
        while True:
            valor_mapa = ws_comparacao.cell(row=linha, column=1).value
            if not valor_mapa or valor_mapa == '':
                break
            if str(valor_mapa) == str(mapa):
                total += 1
                match_val = ws_comparacao.cell(row=linha, column=8).value
                if match_val == 1:
                    matches += 1
            linha += 1

        similaridade = (matches / total) if total > 0 else 0
        mapas_com_similaridade.append((mapa, similaridade))

    mapas_com_similaridade.sort(key=lambda x: x[1], reverse=True)
    print(f"  Mapas ordenados por similaridade: {[(m, f'{s*100:.1f}%') for m, s in mapas_com_similaridade]}")

    # Adiciona dados (uma linha por mapa, já ordenado)
    linha_atual = 4
    for mapa, _ in mapas_com_similaridade:
        # Coluna A: number_mapa
        ws.cell(row=linha_atual, column=1).value = mapa
        ws.cell(row=linha_atual, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=1).border = thin_border

        # Coluna B: Similaridade (%)
        ws.cell(row=linha_atual, column=2).value = (
            f"=IFERROR(SUMIFS('Comparação'!$H$13:$H$5000,'Comparação'!$A$13:$A$5000,A{linha_atual})/"
            f"COUNTIF('Comparação'!$A$13:$A$5000,A{linha_atual}),0)"
        )
        ws.cell(row=linha_atual, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=2).border = thin_border
        ws.cell(row=linha_atual, column=2).number_format = '0.0%'

        # Coluna C: Ocupação Total WMS
        ws.cell(row=linha_atual, column=3).value = (
            f"=SUMIF('Comparação'!$A$13:$A$5000,A{linha_atual},'Comparação'!$G$13:$G$5000)"
        )
        ws.cell(row=linha_atual, column=3).alignment = Alignment(horizontal='right')
        ws.cell(row=linha_atual, column=3).border = thin_border
        ws.cell(row=linha_atual, column=3).number_format = '0.0'

        # Coluna D: Ocupação Total API
        ws.cell(row=linha_atual, column=4).value = (
            f"=SUMIF('Comparação'!$I$13:$I$5000,A{linha_atual},'Comparação'!$O$13:$O$5000)"
        )
        ws.cell(row=linha_atual, column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=linha_atual, column=4).border = thin_border
        ws.cell(row=linha_atual, column=4).number_format = '0.0'

        # Coluna E: Total Itens WMS
        ws.cell(row=linha_atual, column=5).value = (
            f"=SUMIF('Comparação'!$A$13:$A$5000,A{linha_atual},'Comparação'!$E$13:$E$5000)"
        )
        ws.cell(row=linha_atual, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=5).border = thin_border
        ws.cell(row=linha_atual, column=5).number_format = '0'

        # Coluna F: Total Itens API
        ws.cell(row=linha_atual, column=6).value = (
            f"=SUMIF('Comparação'!$I$13:$I$5000,A{linha_atual},'Comparação'!$M$13:$M$5000)"
        )
        ws.cell(row=linha_atual, column=6).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=6).border = thin_border
        ws.cell(row=linha_atual, column=6).number_format = '0'

        # Coluna G: Perc. Não Paletizados WMS
        ws.cell(row=linha_atual, column=7).value = (
            f"=IFERROR(COUNTIF('Não Paletizados'!$A$5:$A$5000,A{linha_atual})/E{linha_atual},0)"
        )
        ws.cell(row=linha_atual, column=7).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=7).border = thin_border
        ws.cell(row=linha_atual, column=7).number_format = '0.0%'

        # Coluna H: Perc. Não Paletizados API
        ws.cell(row=linha_atual, column=8).value = (
            f"=IFERROR(COUNTIF('Não Paletizados'!$H$5:$H$5000,A{linha_atual})/F{linha_atual},0)"
        )
        ws.cell(row=linha_atual, column=8).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=8).border = thin_border
        ws.cell(row=linha_atual, column=8).number_format = '0.0%'

        # Coluna I: Tempo Execução (segs) - valor direto se disponível no summary/log
        tempo = durations_map.get(str(mapa), None)
        if tempo is not None:
            ws.cell(row=linha_atual, column=9).value = float(tempo)
            ws.cell(row=linha_atual, column=9).number_format = '0.00'
        else:
            ws.cell(row=linha_atual, column=9).value = ''
        ws.cell(row=linha_atual, column=9).alignment = Alignment(horizontal='right')
        ws.cell(row=linha_atual, column=9).border = thin_border

        linha_atual += 1

    print(f"  ✅ {len(mapas_com_similaridade)} linhas de dados criadas (linhas 4 a {linha_atual-1})")

    # Cria Table Excel para permitir filtros e ordenação
    from openpyxl.worksheet.table import Table, TableStyleInfo

    # Define range da tabela (sem a linha de totais)
    tab_range = f"A3:I{linha_atual-1}"
    tab = Table(displayName="TabelaConsolidado", ref=tab_range)

    # Estilo da tabela
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style

    ws.add_table(tab)
    print(f"  ✅ Table Excel criada: {tab_range}")

    # Linha de totais (fora da tabela)
    ws.cell(row=linha_atual, column=1).value = 'TOTAL'
    ws.cell(row=linha_atual, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=1).font = Font(bold=True)
    ws.cell(row=linha_atual, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=1).border = thin_border

    # Similaridade TOTAL
    ws.cell(row=linha_atual, column=2).value = (
        f"=IFERROR(SUMPRODUCT(('Comparação'!$A$13:$A$5000<>\"\")*('Comparação'!$H$13:$H$5000=1))/"
        f"SUMPRODUCT(('Comparação'!$A$13:$A$5000<>\"\")* 1),0)"
    )
    ws.cell(row=linha_atual, column=2).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=2).font = Font(bold=True)
    ws.cell(row=linha_atual, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=2).border = thin_border
    ws.cell(row=linha_atual, column=2).number_format = '0.0%'

    # Totais / médias existentes (mantidos)
    ws.cell(row=linha_atual, column=3).value = f'=SUM(C4:C{linha_atual-1})'
    ws.cell(row=linha_atual, column=3).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=3).font = Font(bold=True)
    ws.cell(row=linha_atual, column=3).alignment = Alignment(horizontal='right')
    ws.cell(row=linha_atual, column=3).border = thin_border
    ws.cell(row=linha_atual, column=3).number_format = '0.0'

    ws.cell(row=linha_atual, column=4).value = f'=SUM(D4:D{linha_atual-1})'
    ws.cell(row=linha_atual, column=4).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=4).font = Font(bold=True)
    ws.cell(row=linha_atual, column=4).alignment = Alignment(horizontal='right')
    ws.cell(row=linha_atual, column=4).border = thin_border
    ws.cell(row=linha_atual, column=4).number_format = '0.0'

    ws.cell(row=linha_atual, column=5).value = f'=SUM(E4:E{linha_atual-1})'
    ws.cell(row=linha_atual, column=5).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=5).font = Font(bold=True)
    ws.cell(row=linha_atual, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=5).border = thin_border
    ws.cell(row=linha_atual, column=5).number_format = '0'

    ws.cell(row=linha_atual, column=6).value = f'=SUM(F4:F{linha_atual-1})'
    ws.cell(row=linha_atual, column=6).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=6).font = Font(bold=True)
    ws.cell(row=linha_atual, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=6).border = thin_border
    ws.cell(row=linha_atual, column=6).number_format = '0'

    ws.cell(row=linha_atual, column=7).value = f'=AVERAGE(G4:G{linha_atual-1})'
    ws.cell(row=linha_atual, column=7).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=7).font = Font(bold=True)
    ws.cell(row=linha_atual, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=7).border = thin_border
    ws.cell(row=linha_atual, column=7).number_format = '0.0%'

    ws.cell(row=linha_atual, column=8).value = f'=AVERAGE(H4:H{linha_atual-1})'
    ws.cell(row=linha_atual, column=8).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=8).font = Font(bold=True)
    ws.cell(row=linha_atual, column=8).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=8).border = thin_border
    ws.cell(row=linha_atual, column=8).number_format = '0.0%'

    # Para a coluna de tempo, exibe soma total (opcional) e média
    ws.cell(row=linha_atual, column=9).value = f"=IFERROR(AVERAGE(I4:I{linha_atual-1}),0)"
    ws.cell(row=linha_atual, column=9).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=9).font = Font(bold=True)
    ws.cell(row=linha_atual, column=9).alignment = Alignment(horizontal='right')
    ws.cell(row=linha_atual, column=9).border = thin_border
    ws.cell(row=linha_atual, column=9).number_format = '0.00'

    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 14

    # Freeze panes
    ws.freeze_panes = 'A4'

    print(f"  ✅ Aba Consolidado criada com {len(mapas_unicos)} mapas")
    
def criar_aba_consolidado22(wb, ws_comparacao):
    """
    Cria aba 'Consolidado' com resumo por number_mapa.
    
    Args:
        wb: Workbook
        ws_comparacao: Worksheet da aba 'Comparação'
    """
    from openpyxl.utils import get_column_letter
    
    # Remove aba anterior se existir
    if "Consolidado" in wb.sheetnames:
        del wb["Consolidado"]
    
    ws = wb.create_sheet("Consolidado", 0)  # Cria como primeira aba
    
    # Cores
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = 'Resumo Consolidado por Mapa'
    ws['A1'].fill = header_fill
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].border = thin_border
    
    # Cabeçalhos (linha 3)
    headers = [
        'number_mapa', 
        'Similaridade (%)', 
        'Ocupação Total WMS', 
        'Ocupação Total API',
        'Total Itens WMS',
        'Total Itens API',
        'Perc. Não Paletizados WMS',
        'Perc. Não Paletizados API'
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Extrai number_mapas únicos da aba Comparação
    # Lê coluna A (number_mapa WMS) a partir da linha 13
    mapas_unicos = []
    linha = 13
    while True:
        valor = ws_comparacao.cell(row=linha, column=1).value
        if not valor or valor == '':
            break
        if valor not in mapas_unicos:
            mapas_unicos.append(valor)
        linha += 1
    
    print(f"\n  📊 Criando aba Consolidado com {len(mapas_unicos)} mapas únicos")
    
    # Debug: mostra os mapas encontrados
    print(f"  Mapas únicos encontrados: {mapas_unicos}")
    
    # Calcula similaridade para cada mapa para poder ordenar
    mapas_com_similaridade = []
    for mapa in mapas_unicos:
        # Conta matches e total de linhas para calcular similaridade
        matches = 0
        total = 0
        linha = 13
        while True:
            valor_mapa = ws_comparacao.cell(row=linha, column=1).value
            if not valor_mapa or valor_mapa == '':
                break
            if str(valor_mapa) == str(mapa):
                total += 1
                match_val = ws_comparacao.cell(row=linha, column=8).value
                if match_val == 1:
                    matches += 1
            linha += 1
        
        similaridade = (matches / total) if total > 0 else 0  # Fração 0-1 para ordenação correta
        mapas_com_similaridade.append((mapa, similaridade))
    
    # Ordena por similaridade (do maior para o menor)
    mapas_com_similaridade.sort(key=lambda x: x[1], reverse=True)
    print(f"  Mapas ordenados por similaridade: {[(m, f'{s*100:.1f}%') for m, s in mapas_com_similaridade]}")
    
    # Adiciona dados (uma linha por mapa, já ordenado)
    linha_atual = 4
    for mapa, _ in mapas_com_similaridade:
        # Coluna A: number_mapa
        ws.cell(row=linha_atual, column=1).value = mapa
        ws.cell(row=linha_atual, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=1).border = thin_border
        
        # Coluna B: Similaridade (%)
        # Fórmula: (matches do mapa / total linhas do mapa) * 100
        # Usa aspas simples para nome da aba com acento
        ws.cell(row=linha_atual, column=2).value = (
            f"=IFERROR(SUMIFS('Comparação'!$H$13:$H$5000,'Comparação'!$A$13:$A$5000,A{linha_atual})/"
            f"COUNTIF('Comparação'!$A$13:$A$5000,A{linha_atual}),0)"
        )
        ws.cell(row=linha_atual, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=2).border = thin_border
        ws.cell(row=linha_atual, column=2).number_format = '0.0%'
        
        # Coluna C: Ocupação Total WMS
        # Soma ocupacao (coluna G) do WMS para o mapa
        ws.cell(row=linha_atual, column=3).value = (
            f"=SUMIF('Comparação'!$A$13:$A$5000,A{linha_atual},'Comparação'!$G$13:$G$5000)"
        )
        ws.cell(row=linha_atual, column=3).alignment = Alignment(horizontal='right')
        ws.cell(row=linha_atual, column=3).border = thin_border
        ws.cell(row=linha_atual, column=3).number_format = '0.0'
        
        # Coluna D: Ocupação Total API
        # Soma ocupacao (coluna O) do API para o mapa
        ws.cell(row=linha_atual, column=4).value = (
            f"=SUMIF('Comparação'!$I$13:$I$5000,A{linha_atual},'Comparação'!$O$13:$O$5000)"
        )
        ws.cell(row=linha_atual, column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=linha_atual, column=4).border = thin_border
        ws.cell(row=linha_atual, column=4).number_format = '0.0'
        
        # Coluna E: Total Itens WMS
        # Soma quantidade (quantity) de todos os itens do mapa na aba Comparação (coluna E)
        ws.cell(row=linha_atual, column=5).value = (
            f"=SUMIF('Comparação'!$A$13:$A$5000,A{linha_atual},'Comparação'!$E$13:$E$5000)"
        )
        ws.cell(row=linha_atual, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=5).border = thin_border
        ws.cell(row=linha_atual, column=5).number_format = '0'
        
        # Coluna F: Total Itens API
        # Soma quantidade (quantity) de todos os itens do mapa na aba Comparação (coluna M)
        ws.cell(row=linha_atual, column=6).value = (
            f"=SUMIF('Comparação'!$I$13:$I$5000,A{linha_atual},'Comparação'!$M$13:$M$5000)"
        )
        ws.cell(row=linha_atual, column=6).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=6).border = thin_border
        ws.cell(row=linha_atual, column=6).number_format = '0'
        
        # Coluna G: Perc. Não Paletizados WMS
        # Fórmula: (quantidade de itens não paletizados WMS / total itens WMS) * 100
        # Conta itens (linhas) na aba 'Não Paletizados' para o mapa
        ws.cell(row=linha_atual, column=7).value = (
            f"=IFERROR(COUNTIF('Não Paletizados'!$A$5:$A$5000,A{linha_atual})/E{linha_atual},0)"
        )
        ws.cell(row=linha_atual, column=7).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=7).border = thin_border
        ws.cell(row=linha_atual, column=7).number_format = '0.0%'
        
        # Coluna H: Perc. Não Paletizados API
        # Fórmula: (quantidade de itens não paletizados API / total itens API) * 100
        # Conta itens (linhas) na aba 'Não Paletizados' para o mapa (coluna H é number_mapa API)
        ws.cell(row=linha_atual, column=8).value = (
            f"=IFERROR(COUNTIF('Não Paletizados'!$H$5:$H$5000,A{linha_atual})/F{linha_atual},0)"
        )
        ws.cell(row=linha_atual, column=8).alignment = Alignment(horizontal='center')
        ws.cell(row=linha_atual, column=8).border = thin_border
        ws.cell(row=linha_atual, column=8).number_format = '0.0%'
        
        linha_atual += 1
    
    print(f"  ✅ {len(mapas_com_similaridade)} linhas de dados criadas (linhas 4 a {linha_atual-1})")
    
    # Cria Table Excel para permitir filtros e ordenação
    from openpyxl.worksheet.table import Table, TableStyleInfo
    
    # Define range da tabela (sem a linha de totais)
    tab_range = f"A3:H{linha_atual-1}"
    tab = Table(displayName="TabelaConsolidado", ref=tab_range)
    
    # Estilo da tabela
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    
    ws.add_table(tab)
    print(f"  ✅ Table Excel criada: {tab_range}")
    
    # Linha de totais (fora da tabela)
    ws.cell(row=linha_atual, column=1).value = 'TOTAL'
    ws.cell(row=linha_atual, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=1).font = Font(bold=True)
    ws.cell(row=linha_atual, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=1).border = thin_border
    
    # Similaridade TOTAL (mesma fórmula que aba Comparação G1: total matches / total linhas)
    ws.cell(row=linha_atual, column=2).value = (
        f"=IFERROR(SUMPRODUCT(('Comparação'!$A$13:$A$5000<>\"\")*('Comparação'!$H$13:$H$5000=1))/"
        f"SUMPRODUCT(('Comparação'!$A$13:$A$5000<>\"\")* 1),0)"
    )
    ws.cell(row=linha_atual, column=2).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=2).font = Font(bold=True)
    ws.cell(row=linha_atual, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=2).border = thin_border
    ws.cell(row=linha_atual, column=2).number_format = '0.0%'
    
    # Total ocupação WMS
    ws.cell(row=linha_atual, column=3).value = f'=SUM(C4:C{linha_atual-1})'
    ws.cell(row=linha_atual, column=3).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=3).font = Font(bold=True)
    ws.cell(row=linha_atual, column=3).alignment = Alignment(horizontal='right')
    ws.cell(row=linha_atual, column=3).border = thin_border
    ws.cell(row=linha_atual, column=3).number_format = '0.0'
    
    # Total ocupação API
    ws.cell(row=linha_atual, column=4).value = f'=SUM(D4:D{linha_atual-1})'
    ws.cell(row=linha_atual, column=4).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=4).font = Font(bold=True)
    ws.cell(row=linha_atual, column=4).alignment = Alignment(horizontal='right')
    ws.cell(row=linha_atual, column=4).border = thin_border
    ws.cell(row=linha_atual, column=4).number_format = '0.0'
    
    # Total itens WMS
    ws.cell(row=linha_atual, column=5).value = f'=SUM(E4:E{linha_atual-1})'
    ws.cell(row=linha_atual, column=5).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=5).font = Font(bold=True)
    ws.cell(row=linha_atual, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=5).border = thin_border
    ws.cell(row=linha_atual, column=5).number_format = '0'
    
    # Total itens API
    ws.cell(row=linha_atual, column=6).value = f'=SUM(F4:F{linha_atual-1})'
    ws.cell(row=linha_atual, column=6).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=6).font = Font(bold=True)
    ws.cell(row=linha_atual, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=6).border = thin_border
    ws.cell(row=linha_atual, column=6).number_format = '0'
    
    # Média percentual não paletizados WMS
    ws.cell(row=linha_atual, column=7).value = f'=AVERAGE(G4:G{linha_atual-1})'
    ws.cell(row=linha_atual, column=7).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=7).font = Font(bold=True)
    ws.cell(row=linha_atual, column=7).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=7).border = thin_border
    ws.cell(row=linha_atual, column=7).number_format = '0.0%'
    
    # Média percentual não paletizados API
    ws.cell(row=linha_atual, column=8).value = f'=AVERAGE(H4:H{linha_atual-1})'
    ws.cell(row=linha_atual, column=8).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws.cell(row=linha_atual, column=8).font = Font(bold=True)
    ws.cell(row=linha_atual, column=8).alignment = Alignment(horizontal='center')
    ws.cell(row=linha_atual, column=8).border = thin_border
    ws.cell(row=linha_atual, column=8).number_format = '0.0%'
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 25
    
    # Freeze panes
    ws.freeze_panes = 'A4'
    
    print(f"  ✅ Aba Consolidado criada com {len(mapas_unicos)} mapas")


def criar_aba_nao_paletizados(wb, ws_comparacao):
    """
    Cria aba 'Não Paletizados' comparando produtos fora do caminhão entre WMS e API.
    Processa TODOS os mapas presentes na aba de Comparação.
    
    Args:
        wb: Workbook
        ws_comparacao: Worksheet da aba 'Comparação' com todos os mapas
    """
    # Remove aba anterior se existir
    if "Não Paletizados" in wb.sheetnames:
        del wb["Não Paletizados"]
    
    ws = wb.create_sheet("Não Paletizados")
    
    # Cores
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    separador_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    print(f"\n  [INFO] Extraindo produtos nao paletizados de TODOS os mapas processados")
    
    # Extrai todos os number_mapa únicos da aba Comparação (coluna A, a partir da linha 13)
    mapas_unicos = set()
    for row in range(13, ws_comparacao.max_row + 1):
        mapa = ws_comparacao.cell(row=row, column=1).value
        if mapa:
            mapas_unicos.add(str(mapa).strip())
    
    print(f"  [INFO] Encontrados {len(mapas_unicos)} mapas unicos: {sorted(mapas_unicos)}")
    
    # Para cada mapa, encontra os arquivos e extrai produtos não paletizados
    produtos_wms_todos = []
    produtos_api_todos = []
    
    # Usa o mapeamento global de arquivos processados
    global _MAPEAMENTO_ARQUIVOS
    print(f"  [INFO] Mapeamento global tem {len(_MAPEAMENTO_ARQUIVOS)} entradas")
    print(f"  [DEBUG] Chaves disponiveis: {list(_MAPEAMENTO_ARQUIVOS.keys())}")
    
    # Descobre arquivos para mapas que NÃO estão no mapeamento
    mapas_faltantes = [m for m in mapas_unicos if m not in _MAPEAMENTO_ARQUIVOS]
    
    if len(mapas_faltantes) > 0:
        print(f"  [WARN] {len(mapas_faltantes)} mapas faltando no mapeamento: {mapas_faltantes}")
        print(f"  [INFO] Tentando descobrir arquivos pela estrutura de pastas...")
        
        # Tenta descobrir o diretório base a partir de múltiplos locais possíveis
        base_paths = [
            Path.cwd(),  # Diretório atual (pode ser onde o Excel está)
            # Path(r'C:\Users\BRKEY864393\Downloads\routes_2k\routes'),  # Caminho batch comum
            Path(r'C:\Users\BRKEY864393\Downloads\route_em_massa'),  # Caminho batch comum
            # Path(r'C:\Users\BRKEY864393\Downloads\execucao_em_massa'),
            Path.cwd() / 'data' / 'route',
            Path(__file__).parent.parent / 'data' / 'route',
        ]
        
        print(f"  [DEBUG] Caminhos base a tentar:")
        for bp in base_paths:
            print(f"     - {bp}")
        
        for mapa in mapas_faltantes:
            wms_file = None
            api_file = None
            
            for base_path in base_paths:
                map_folder = base_path / mapa
                if map_folder.exists():
                    wms_candidate = map_folder / 'output.txt'
                    if wms_candidate.exists():
                        wms_file = wms_candidate
                    else:
                        # 2) fallback: qualquer arquivo que termine com 'ocp-Rota.txt'
                        rota_candidates = list(map_folder.glob('*ocp-Rota.txt'))
                        if rota_candidates:
                            wms_file = rota_candidates[0]
                    if wms_candidate.exists():
                        wms_file = wms_candidate
                        
                        output_dir = map_folder / 'output'
                        if output_dir.exists():
                            api_files = list(output_dir.glob(f'palletize_result_map_{mapa}.txt'))
                            if not api_files:
                                api_files = list(output_dir.glob('palletize_result_map_*.txt'))
                            if api_files:
                                api_file = api_files[0]
                                break
                
                if wms_file and api_file:
                    break
            
            if wms_file and api_file:
                _MAPEAMENTO_ARQUIVOS[mapa] = (wms_file, api_file)
                print(f"     [OK] Mapa {mapa} descoberto: {wms_file.parent.parent}")
            else:
                print(f"     [ERRO] Mapa {mapa} NÃO foi encontrado em nenhum caminho!")
        
        print(f"  [INFO] Após descoberta: {len(_MAPEAMENTO_ARQUIVOS)} mapas no mapeamento total")
    
    for mapa in sorted(mapas_unicos):
        print(f"\n  [INFO] Processando mapa {mapa}...")
        
        # Tenta obter arquivos do mapeamento global
        if mapa in _MAPEAMENTO_ARQUIVOS:
            wms_file, api_file = _MAPEAMENTO_ARQUIVOS[mapa]
            print(f"     [OK] Arquivos encontrados no mapeamento global")
            print(f"     WMS: {wms_file}")
            print(f"     API: {api_file}")
        else:
            print(f"     [WARN] Mapa {mapa} nao encontrado no mapeamento global")
            wms_file = None
            api_file = None
        
        if wms_file and api_file and wms_file.exists() and api_file.exists():
            
            prods_wms = extrair_produtos_nao_paletizados(wms_file, number_mapa=mapa)
            prods_api = extrair_produtos_nao_paletizados(api_file, number_mapa=mapa)
            
            produtos_wms_todos.extend(prods_wms)
            produtos_api_todos.extend(prods_api)
            
            print(f"     [OK] {len(prods_wms)} produtos WMS, {len(prods_api)} produtos API")
        else:
            print(f"     [WARN] Arquivos nao encontrados para mapa {mapa}")
    
    print(f"\n  [INFO] Total agregado: {len(produtos_wms_todos)} produtos WMS, {len(produtos_api_todos)} produtos API")
    
    # Título
    ws.merge_cells('A1:N1')
    ws['A1'] = 'Produtos Não Paletizados (Fora do Caminhão)'
    ws['A1'].fill = header_fill
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Cabeçalhos principais (linha 3)
    ws.merge_cells('A3:F3')
    ws['A3'] = 'WMS'
    ws['A3'].fill = header_fill
    ws['A3'].font = header_font
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['G3'].fill = separador_fill
    
    ws.merge_cells('H3:M3')
    ws['H3'] = 'RESULTADO'
    ws['H3'].fill = header_fill
    ws['H3'].font = header_font
    ws['H3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['N3'].fill = separador_fill
    
    # Cabeçalhos de colunas (linha 4)
    headers_wms = ['number_mapa', 'product_code', 'product_name', 'quantity', 'atributo', 'ocupacao']
    for col, header in enumerate(headers_wms, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.cell(row=4, column=7).fill = separador_fill
    
    headers_api = ['number_mapa', 'product_code', 'product_name', 'quantity', 'atributo', 'ocupacao']
    for col, header in enumerate(headers_api, start=8):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.cell(row=4, column=14).fill = separador_fill
    ws.cell(row=4, column=14).value = 'Match'
    ws.cell(row=4, column=14).font = Font(size=8, color="666666")
    
    # Define colunas esperadas ANTES de criar DataFrames
    colunas_esperadas = ['number_mapa', 'product_code', 'product_name', 'quantity', 'atributo', 'ocupacao']
    
    # Converte para DataFrames com colunas garantidas
    if produtos_wms_todos:
        df_wms = pd.DataFrame(produtos_wms_todos)
        # Garante que todas as colunas existam
        for col in colunas_esperadas:
            if col not in df_wms.columns:
                df_wms[col] = ''
        df_wms = df_wms[colunas_esperadas]
        # Converte number_mapa para string para evitar NaN/vazios
        if 'number_mapa' in df_wms.columns:
            df_wms['number_mapa'] = df_wms['number_mapa'].astype(str)
    else:
        df_wms = pd.DataFrame(columns=colunas_esperadas)
    
    if produtos_api_todos:
        df_api = pd.DataFrame(produtos_api_todos)
        # Garante que todas as colunas existam
        for col in colunas_esperadas:
            if col not in df_api.columns:
                df_api[col] = ''
        df_api = df_api[colunas_esperadas]
        # Converte number_mapa para string para evitar NaN/vazios
        if 'number_mapa' in df_api.columns:
            df_api['number_mapa'] = df_api['number_mapa'].astype(str)
    else:
        df_api = pd.DataFrame(columns=colunas_esperadas)
    
    print(f"  [INFO] DataFrames criados:")
    print(f"     WMS shape: {df_wms.shape}")
    print(f"     API shape: {df_api.shape}")
    if len(df_wms) > 0:
        print(f"     WMS colunas: {list(df_wms.columns)}")
        print(f"     WMS primeiras 3 linhas:\n{df_wms.head(3)}")
    if len(df_api) > 0:
        print(f"     API colunas: {list(df_api.columns)}")
        print(f"     API primeiras 3 linhas:\n{df_api.head(3)}")
    
    # Se ambos estiverem vazios, avisa mas mantém a aba para debug
    if len(df_wms) == 0 and len(df_api) == 0:
        print(f"  [WARN] ⚠️ Nenhum produto nao paletizado encontrado!")
        print(f"  [INFO] Possíveis causas:")
        print(f"     1. Mapas não têm produtos fora do caminhão")
        print(f"     2. Mapeamento global está vazio (processar mapas antes de criar Excel)")
        print(f"     3. Arquivos não foram encontrados nos caminhos mapeados")
        print(f"  [INFO] Mantendo aba vazia para análise...")
        # NÃO deleta a aba - mantém para debug
        # del wb["Não Paletizados"]
        # return
    
    print(f"  [INFO] Continuando com criacao da aba (tem dados)")
    
    # Equaliza número de linhas
    max_linhas = max(len(df_wms), len(df_api), 1)  # Pelo menos 1 linha
    
    print(f"  [INFO] Equalizando para {max_linhas} linhas")
    
    if len(df_wms) < max_linhas:
        df_wms_vazio = pd.DataFrame('', index=range(max_linhas - len(df_wms)), columns=colunas_esperadas)
        df_wms = pd.concat([df_wms, df_wms_vazio], ignore_index=True)
    
    if len(df_api) < max_linhas:
        df_api_vazio = pd.DataFrame('', index=range(max_linhas - len(df_api)), columns=colunas_esperadas)
        df_api = pd.concat([df_api, df_api_vazio], ignore_index=True)
    
    # Preenche células vazias com string vazia, exceto number_mapa que já foi convertido
    for col in colunas_esperadas:
        if col != 'number_mapa':  # Preserva number_mapa
            df_wms[col] = df_wms[col].fillna('')
            df_api[col] = df_api[col].fillna('')
    
    print(f"  [INFO] Escrevendo {max_linhas} linhas na planilha...")
    
    # DEBUG: Mostra os dados antes de escrever
    print(f"\n  [DEBUG] Dados a serem escritos:")
    if len(df_wms) > 0:
        print(f"     WMS linha 0: {df_wms.iloc[0].to_dict()}")
    if len(df_api) > 0:
        print(f"     API linha 0: {df_api.iloc[0].to_dict()}")
    
    # Escreve dados
    linha_atual = 5
    try:
        for i in range(max_linhas):
            # WMS - colunas 1 a 6
            for col_idx, col_name in enumerate(colunas_esperadas, start=1):
                valor = df_wms.iloc[i][col_name] if i < len(df_wms) else ''
                ws.cell(row=linha_atual, column=col_idx).value = valor
                ws.cell(row=linha_atual, column=col_idx).border = thin_border
                # DEBUG primeira linha
                if i == 0:
                    print(f"     WMS col {col_idx} ({col_name}): '{valor}'")
            
            # Separador (coluna 7)
            ws.cell(row=linha_atual, column=7).fill = separador_fill
            
            # API - colunas 8 a 13 (6 colunas)
            for idx, col_name in enumerate(colunas_esperadas):
                col_idx = idx + 8  # Começa na coluna 8
                valor = df_api.iloc[i][col_name] if i < len(df_api) else ''
                ws.cell(row=linha_atual, column=col_idx).value = valor
                ws.cell(row=linha_atual, column=col_idx).border = thin_border
                # DEBUG primeira linha
                if i == 0:
                    print(f"     API col {col_idx} ({col_name}): '{valor}'")
            
            # Match (coluna N = 14)
            ws.cell(row=linha_atual, column=14).value = (
                f"=IF(TRIM(A{linha_atual})<>\"\",IF(COUNTIFS($H$5:$H$5000,TRIM(A{linha_atual}),"
                f"$I$5:$I$5000,TRIM(B{linha_atual}),$K$5:$K$5000,D{linha_atual})>0,1,\"\"),\"\")"
            )
            ws.cell(row=linha_atual, column=14).fill = separador_fill
            ws.cell(row=linha_atual, column=14).font = Font(size=8)
            
            linha_atual += 1
    except Exception as e:
        print(f"  [ERRO] Erro ao escrever linha {i}: {e}")
        import traceback
        traceback.print_exc()
    
    print(f"  [INFO] {linha_atual - 5} linhas foram escritas no worksheet")
    
    # Ativa AutoFilter na linha de cabeçalho para permitir filtros
    ws.auto_filter.ref = f'A4:N{max_linhas + 4}'
    print(f"  [INFO] AutoFilter configurado: A4:N{max_linhas + 4}")
    
    # Ajusta larguras
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 2
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 50
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 10
    ws.column_dimensions['N'].width = 2
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    print(f"  [OK] Aba Nao Paletizados criada: {len(df_wms)} linhas WMS, {len(df_api)} linhas API")


def processar_mapa_individual(map_folder: Path, output_file: Path, append: bool = False) -> bool:
    """
    Processa um mapa individual e adiciona à planilha consolidada.
    
    Args:
        map_folder: Pasta do mapa contendo output.txt e output/palletize_result_map_*.txt
        output_file: Arquivo Excel consolidado
        append: Se True, concatena com dados existentes
        
    Returns:
        True se processado com sucesso, False caso contrário
    """
    # Encontra arquivo WMS (output.txt)
    wms_file = map_folder / 'output.txt'
    if not wms_file.exists():
        print(f"⚠️ Arquivo WMS não encontrado: {wms_file}")
        return False
    
    # Encontra arquivo API (palletize_result_map_*.txt)
    output_dir = map_folder / 'output'
    if not output_dir.exists():
        print(f"⚠️ Pasta output não encontrada: {output_dir}")
        return False
    
    # Procura arquivo palletize_result_map_*.txt
    api_files = list(output_dir.glob('palletize_result_map_*.txt'))
    if not api_files:
        print(f"⚠️ Arquivo API não encontrado em: {output_dir}")
        return False
    
    api_file = api_files[0]  # Usa o primeiro encontrado
    
    try:
        criar_excel_comparacao(wms_file, api_file, output_file, append=append)
        return True
    except Exception as e:
        print(f"❌ Erro ao processar mapa {map_folder.name}: {e}")
        import traceback
        traceback.print_exc()
        return False


def processar_batch_mapas(root_folder: Path) -> Path:
    """
    Processa todos os mapas em uma pasta raiz e gera planilha consolidada.
    
    Args:
        root_folder: Pasta raiz contendo subpastas com mapas
        
    Returns:
        Caminho do arquivo Excel consolidado
    """
    print("\n" + "=" * 80)
    print("PROCESSAMENTO EM LOTE - COMPARAÇÃO WMS vs API")
    print("=" * 80)
    print(f"Pasta raiz: {root_folder}\n")
    
    # Arquivo de saída consolidado
    output_file = root_folder / 'comparacao_consolidada_wms_vs_api.xlsx'
    
    # Remove arquivo existente para começar do zero
    if output_file.exists():
        output_file.unlink()
        print(f"♻️ Arquivo consolidado anterior removido\n")
    
    # Encontra todas as pastas de mapas
    map_folders = sorted([f for f in root_folder.iterdir() if f.is_dir()])
    
    if not map_folders:
        print(f"⚠️ Nenhuma pasta de mapa encontrada em: {root_folder}")
        return None
    
    print(f"📁 Total de pastas encontradas: {len(map_folders)}\n")
    
    processed_count = 0
    skipped_count = 0
    
    for i, map_folder in enumerate(map_folders, 1):
        print(f"\n[{i}/{len(map_folders)}] Processando: {map_folder.name}")
        print("-" * 60)
        
        # Processa mapa (append=True para concatenar após o primeiro)
        success = processar_mapa_individual(
            map_folder=map_folder,
            output_file=output_file,
            append=(processed_count > 0)  # Concatena a partir do 2º mapa
        )
        
        if success:
            processed_count += 1
            print(f"✅ Mapa {map_folder.name} adicionado à planilha consolidada")
        else:
            skipped_count += 1
            print(f"⏭️ Mapa {map_folder.name} ignorado")
    
    # Após processar TODOS os mapas, cria a aba "Não Paletizados" com dados agregados
    if processed_count > 0 and output_file.exists():
        print("\n" + "=" * 80)
        print("CRIANDO ABA 'NÃO PALETIZADOS' COM TODOS OS MAPAS")
        print("=" * 80)
        
        # Debug: verifica estado do mapeamento global
        global _MAPEAMENTO_ARQUIVOS
        print(f"[DEBUG] Mapeamento global no final do batch:")
        print(f"        Tamanho: {len(_MAPEAMENTO_ARQUIVOS)}")
        print(f"        Chaves: {list(_MAPEAMENTO_ARQUIVOS.keys())}")
        
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        ws_comparacao = wb['Comparação']
        
        # Cria aba com produtos não paletizados de TODOS os mapas
        criar_aba_nao_paletizados(wb, ws_comparacao)
        
        wb.save(output_file)
        wb.close()
        print("✅ Aba 'Não Paletizados' criada com sucesso!")
        print("=" * 80)
    
    print("\n" + "=" * 80)
    print("RESUMO DO PROCESSAMENTO")
    print("=" * 80)
    print(f"✅ Mapas processados: {processed_count}")
    print(f"⏭️ Mapas ignorados: {skipped_count}")
    print(f"📊 Total de pastas: {len(map_folders)}")
    print(f"\n📄 Planilha consolidada: {output_file}")
    print("=" * 80)
    
    return output_file


if __name__ == '__main__':
    # Caminhos dos arquivos
    base_dir = Path(__file__).parent.parent / 'data' / 'route' / '620768'
    
    wms_file = base_dir / 'output.txt'
    api_file = base_dir / 'output' / 'palletize_result_map_620768.txt'
    output_file = base_dir / 'output' / 'comparacao_wms_vs_api.xlsx'
    
    # Verifica se arquivos existem
    if not wms_file.exists():
        print(f"ERRO: Arquivo WMS não encontrado: {wms_file}")
        exit(1)
    
    if not api_file.exists():
        print(f"ERRO: Arquivo API não encontrado: {api_file}")
        exit(1)
    
    # Cria comparação
    criar_excel_comparacao(wms_file, api_file, output_file)
