"""
Script para gerar relatório comparativo entre outputs do C# e Python.

Lê dois arquivos .txt (output C# e Python) e gera um arquivo Excel com:
- KPIs de similaridade (pallets, produtos, quantidades)
- Tabela pivot por código de pallet e produto
- Comparação lado a lado dos pallets
"""
import argparse
import re
from pathlib import Path
from typing import Dict, List, Tuple
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class PalletProduct:
    """Representa um produto em um pallet."""
    def __init__(self, code: int, name: str, quantity: int, unit: int):
        self.code = code
        self.name = name
        self.quantity = quantity
        self.unit = unit


class Pallet:
    """Representa um pallet com seus produtos."""
    def __init__(self, pallet_id: str, side: str, occupation: float, delivery: str, weight: float):
        self.pallet_id = pallet_id
        self.side = side
        self.occupation = occupation
        self.delivery = delivery
        self.weight = weight
        self.products: List[PalletProduct] = []


class OutputParser:
    """Parser para arquivos de output do stackbuilder."""
    
    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.content = self._read_file()
        
    def _read_file(self) -> str:
        """Lê o arquivo com encoding adequado."""
        encodings = ['utf-8', 'latin-1', 'cp1252']
        for encoding in encodings:
            try:
                return self.file_path.read_text(encoding=encoding)
            except UnicodeDecodeError:
                continue
        raise ValueError(f"Não foi possível ler o arquivo {self.file_path}")
    
    def parse(self) -> Tuple[Dict[str, Pallet], Dict[str, int]]:
        """
        Parse do arquivo de output.
        
        Returns:
            Tuple com:
            - Dict de pallets {pallet_id: Pallet}
            - Dict de produtos fora do caminhão {code: quantity}
        """
        pallets = {}
        outside_products = {}
        
        # Regex patterns - captura o código completo do pallet
        # Aceita tanto P01_M_01_1/35 quanto P_M_01_1/35
        pallet_pattern = r'P(?:\d+)?_([AM])_(\d+)_(\d+/\d+)\s*-\s*([\d,\.]+)\s*-\s*([^\s]+)\s*-\s*\d+\s+Peso:\s*([\d,\.]+)'
        # Pattern corrigido: sequencia, código, nome, quantidade, embalagem (4 dígitos), grupo/sub
        # A embalagem é SEMPRE 4 dígitos exatos, a quantidade vem imediatamente antes
        product_pattern = r'\|\s*\d+\s+(\d+)\s+(.+?)\s+(\d+)\s+(\d{4})\s+\d+/\d+'
        outside_pattern = r'\|\s*(\d+)\s+(.+?)\s+(\d+)\s+(\d{4})\s+\d+/\d+'
        
        lines = self.content.split('\n')
        current_pallet = None
        in_outside_section = False
        
        for line in lines:
            # Detectar seção de produtos fora
            if 'Produtos fora do caminhão' in line:
                in_outside_section = True
                continue
            
            # Parse de pallet header
            pallet_match = re.search(pallet_pattern, line)
            if pallet_match:
                side = pallet_match.group(1)
                number = pallet_match.group(2)
                sequence = pallet_match.group(3)
                occupation = float(pallet_match.group(4).replace(',', '.'))
                delivery = pallet_match.group(5)
                weight = float(pallet_match.group(6).replace(',', '.'))
                
                # Formato padronizado: P01_M_01_1/35 (com zero-padding)
                number_padded = number.zfill(2)  # Adiciona zero à esquerda se necessário
                pallet_id = f"P{number_padded}_{side}_{number}_{sequence}"
                current_pallet = Pallet(pallet_id, side, occupation, delivery, weight)
                pallets[pallet_id] = current_pallet
                in_outside_section = False
                continue
            
            # Parse de produtos (dentro ou fora)
            if '|' in line and re.search(r'\d+\s+\d+\s+', line):
                if in_outside_section:
                    match = re.search(outside_pattern, line)
                    if match:
                        code = int(match.group(1))
                        quantity = int(match.group(3))
                        outside_products[code] = outside_products.get(code, 0) + quantity
                elif current_pallet:
                    match = re.search(product_pattern, line)
                    if match:
                        code = int(match.group(1))
                        name = match.group(2).strip()
                        quantity = int(match.group(3))
                        unit = int(match.group(4))
                        
                        product = PalletProduct(code, name, quantity, unit)
                        current_pallet.products.append(product)
        
        return pallets, outside_products


class ReportGenerator:
    """Gerador de relatório Excel comparativo."""
    
    def __init__(self, csharp_file: Path, python_file: Path, output_file: Path):
        self.csharp_file = csharp_file
        self.python_file = python_file
        self.output_file = output_file
        
        # Extrair código da rota do caminho do arquivo (ex: route\620768\)
        self.route_code = self._extract_route_code(csharp_file)
        
        # Parse dos arquivos
        self.csharp_pallets, self.csharp_outside = OutputParser(csharp_file).parse()
        self.python_pallets, self.python_outside = OutputParser(python_file).parse()
    
    def _extract_route_code(self, file_path: Path) -> str:
        """Extrai o código da rota do caminho do arquivo."""
        # Procurar por 'route\' ou 'route/' no caminho
        path_str = str(file_path)
        
        # Tentar encontrar padrão route\NUMERO ou route/NUMERO
        import re
        match = re.search(r'route[/\\](\d+)', path_str, re.IGNORECASE)
        if match:
            return match.group(1)
        
        # Fallback: usar timestamp se não encontrar
        from datetime import datetime
        return datetime.now().strftime('%Y%m%d_%H%M%S')
        
    def _apply_header_style(self, cell):
        """Aplica estilo de cabeçalho."""
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _apply_cell_style(self, cell, bg_color=None):
        """Aplica estilo de célula."""
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    
    def _calculate_kpis(self) -> Dict[str, float]:
        """Calcula KPIs de similaridade."""
        kpis = {
            'total_pallets_csharp': len(self.csharp_pallets),
            'total_pallets_python': len(self.python_pallets),
            'pallets_match': 0,
            'products_match': 0,
            'quantities_match': 0,
            'total_products_csharp': 0,
            'total_products_python': 0,
            'outside_products_wms': sum(self.csharp_outside.values()),
            'outside_products_api': sum(self.python_outside.values()),
        }
        
        # Contar produtos totais
        csharp_products = {}
        python_products = {}
        
        for pallet in self.csharp_pallets.values():
            for prod in pallet.products:
                key = f"{pallet.pallet_id}_{prod.code}"
                csharp_products[key] = prod.quantity
                kpis['total_products_csharp'] += 1
        
        for pallet in self.python_pallets.values():
            for prod in pallet.products:
                key = f"{pallet.pallet_id}_{prod.code}"
                python_products[key] = prod.quantity
                kpis['total_products_python'] += 1
        
        # Pallets que existem em ambos
        common_pallets = set(self.csharp_pallets.keys()) & set(self.python_pallets.keys())
        kpis['pallets_match'] = len(common_pallets)
        
        # Produtos que existem em ambos no mesmo pallet
        common_products = set(csharp_products.keys()) & set(python_products.keys())
        kpis['products_match'] = len(common_products)
        
        # Produtos com mesma quantidade
        for key in common_products:
            if csharp_products[key] == python_products[key]:
                kpis['quantities_match'] += 1
        
        # Calcular percentuais
        kpis['pallet_similarity'] = (kpis['pallets_match'] / max(kpis['total_pallets_csharp'], 1)) * 100
        kpis['product_similarity'] = (kpis['products_match'] / max(kpis['total_products_csharp'], 1)) * 100
        kpis['quantity_similarity'] = (kpis['quantities_match'] / max(kpis['products_match'], 1)) * 100 if kpis['products_match'] > 0 else 0
        
        return kpis
    
    def _write_kpis(self, ws):
        """Escreve KPIs na planilha."""
        kpis = self._calculate_kpis()
        
        # Legenda explicativa
        ws['B1'] = 'LEGENDA:'
        ws['B1'].font = Font(bold=True, size=10, color="366092")
        ws['C1'] = 'WMS = Resultado esperado (validação C#) | API = Resultado obtido (Python) | Match = Itens que aparecem em ambos | Similaridade = % de correspondência'
        ws['C1'].font = Font(size=9, italic=True)
        ws.merge_cells('C1:F1')
        
        # Título
        ws['C2'] = 'KPIs - Similaridade'
        ws['C2'].font = Font(bold=True, size=14, color="366092")
        
        # Headers
        row = 4
        ws[f'B{row}'] = 'Métrica'
        ws[f'C{row}'] = 'WMS'
        ws[f'D{row}'] = 'API'
        ws[f'E{row}'] = 'Match'
        ws[f'F{row}'] = 'Similaridade %'
        
        for col in ['B', 'C', 'D', 'E', 'F']:
            self._apply_header_style(ws[f'{col}{row}'])
        
        # Dados
        row += 1
        ws[f'B{row}'] = 'Total Pallets'
        ws[f'C{row}'] = kpis['total_pallets_csharp']
        ws[f'D{row}'] = kpis['total_pallets_python']
        ws[f'E{row}'] = kpis['pallets_match']
        ws[f'F{row}'] = f"{kpis['pallet_similarity']:.2f}%"
        
        row += 1
        ws[f'B{row}'] = 'Total Produtos'
        ws[f'C{row}'] = kpis['total_products_csharp']
        ws[f'D{row}'] = kpis['total_products_python']
        ws[f'E{row}'] = kpis['products_match']
        ws[f'F{row}'] = f"{kpis['product_similarity']:.2f}%"
        
        row += 1
        ws[f'B{row}'] = 'Quantidades Idênticas'
        ws[f'C{row}'] = '-'
        ws[f'D{row}'] = '-'
        ws[f'E{row}'] = kpis['quantities_match']
        ws[f'F{row}'] = f"{kpis['quantity_similarity']:.2f}%"
        
        row += 1
        ws[f'B{row}'] = 'Produtos Não Paletizados'
        ws[f'C{row}'] = kpis['outside_products_wms']
        ws[f'D{row}'] = kpis['outside_products_api']
        diff = kpis['outside_products_api'] - kpis['outside_products_wms']
        ws[f'E{row}'] = diff
        
        # Status: menos é melhor, igual é bom, mais é ruim
        if diff < 0:
            ws[f'F{row}'] = 'Melhor ✓'
            bg_color = 'C6EFCE'
        elif diff == 0:
            ws[f'F{row}'] = 'Igual'
            bg_color = 'FFEB9C'
        else:
            ws[f'F{row}'] = 'Pior ✗'
            bg_color = 'FFC7CE'
        
        ws[f'F{row}'].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        # Aplicar estilos
        for r in range(5, 9):
            for col in ['B', 'C', 'D', 'E', 'F']:
                self._apply_cell_style(ws[f'{col}{r}'])
        
        # Ajustar larguras
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
    
    def _write_product_type_comparison(self, ws):
        """Escreve comparação por tipo de produto."""
        row = 10
        
        # Título
        ws['K10'] = 'Comparação por Tipo de Produto'
        ws['K10'].font = Font(bold=True, size=12, color="366092")
        ws.merge_cells('K10:T10')
        
        row = 12
        
        # Headers principais
        ws['K11'] = 'WMS'
        ws['K11'].font = Font(bold=True, size=11, color="366092")
        ws.merge_cells('K11:P11')
        ws['K11'].alignment = Alignment(horizontal="center", vertical="center")
        
        ws['R11'] = 'RESULTADO'
        ws['R11'].font = Font(bold=True, size=11, color="366092")
        ws.merge_cells('R11:W11')
        ws['R11'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers colunas - WMS
        headers_wms = ['Pallet', 'Descartável', 'Isotonico', 'Retornavel', 'TopoPallet', 'ALL']
        for i, header in enumerate(headers_wms):
            col = chr(ord('K') + i)
            ws[f'{col}{row}'] = header
            self._apply_header_style(ws[f'{col}{row}'])
        
        # Coluna de separação (Q) - vazia
        ws[f'Q{row}'] = ''
        ws[f'Q{row}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Headers colunas - RESULTADO
        headers_api = ['Pallet', 'Descartável', 'Isotonico', 'Retornavel', 'TopoPallet', 'ALL']
        for i, header in enumerate(headers_api):
            col = chr(ord('R') + i)
            ws[f'{col}{row}'] = header
            self._apply_header_style(ws[f'{col}{row}'])
        
        row += 1
        start_row = row
        
        # Coletar dados por pallet e tipo
        wms_data = {}
        api_data = {}
        
        # Processar WMS
        for pallet_id, pallet in self.csharp_pallets.items():
            if pallet_id not in wms_data:
                wms_data[pallet_id] = {'Descartável': 0, 'Isotonico': 0, 'Retornavel': 0, 'TopoPallet': 0}
            for prod in pallet.products:
                prod_type = self._guess_product_type(prod.name)
                if prod_type in wms_data[pallet_id]:
                    wms_data[pallet_id][prod_type] += prod.quantity
        
        # Processar API
        for pallet_id, pallet in self.python_pallets.items():
            if pallet_id not in api_data:
                api_data[pallet_id] = {'Descartável': 0, 'Isotonico': 0, 'Retornavel': 0, 'TopoPallet': 0}
            for prod in pallet.products:
                prod_type = self._guess_product_type(prod.name)
                if prod_type in api_data[pallet_id]:
                    api_data[pallet_id][prod_type] += prod.quantity
        
        # Totais
        wms_totals = {'Descartável': 0, 'Isotonico': 0, 'Retornavel': 0, 'TopoPallet': 0}
        api_totals = {'Descartável': 0, 'Isotonico': 0, 'Retornavel': 0, 'TopoPallet': 0}
        
        # Escrever dados
        all_pallets = sorted(set(list(wms_data.keys()) + list(api_data.keys())))
        
        for pallet_id in all_pallets:
            wms_row = wms_data.get(pallet_id, {'Descartável': 0, 'Isotonico': 0, 'Retornavel': 0, 'TopoPallet': 0})
            api_row = api_data.get(pallet_id, {'Descartável': 0, 'Isotonico': 0, 'Retornavel': 0, 'TopoPallet': 0})
            
            # WMS
            ws[f'K{row}'] = pallet_id
            ws[f'L{row}'] = wms_row['Descartável']
            ws[f'M{row}'] = wms_row['Isotonico']
            ws[f'N{row}'] = wms_row['Retornavel']
            ws[f'O{row}'] = wms_row['TopoPallet']
            ws[f'P{row}'] = sum(wms_row.values())
            
            # Coluna de separação (Q) - vazia
            ws[f'Q{row}'] = ''
            
            # RESULTADO
            ws[f'R{row}'] = pallet_id
            ws[f'S{row}'] = api_row['Descartável']
            ws[f'T{row}'] = api_row['Isotonico']
            ws[f'U{row}'] = api_row['Retornavel']
            ws[f'V{row}'] = api_row['TopoPallet']
            ws[f'W{row}'] = sum(api_row.values())
            
            # Aplicar estilos
            for col in ['K', 'L', 'M', 'N', 'O', 'P']:
                self._apply_cell_style(ws[f'{col}{row}'])
            
            for col in ['R', 'S', 'T', 'U', 'V', 'W']:
                self._apply_cell_style(ws[f'{col}{row}'])
            
            # Acumular totais
            for key in wms_totals:
                wms_totals[key] += wms_row[key]
                api_totals[key] += api_row[key]
            
            row += 1
        
        # Linha de totais
        ws[f'K{row}'] = 'ALL'
        ws[f'L{row}'] = wms_totals['Descartável']
        ws[f'M{row}'] = wms_totals['Isotonico']
        ws[f'N{row}'] = wms_totals['Retornavel']
        ws[f'O{row}'] = wms_totals['TopoPallet']
        ws[f'P{row}'] = sum(wms_totals.values())
        
        ws[f'Q{row}'] = ''
        
        ws[f'R{row}'] = 'ALL'
        ws[f'S{row}'] = api_totals['Descartável']
        ws[f'T{row}'] = api_totals['Isotonico']
        ws[f'U{row}'] = api_totals['Retornavel']
        ws[f'V{row}'] = api_totals['TopoPallet']
        ws[f'W{row}'] = sum(api_totals.values())
        
        # Aplicar estilos de totais (negrito)
        for col in ['K', 'L', 'M', 'N', 'O', 'P']:
            cell = ws[f'{col}{row}']
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            self._apply_cell_style(cell)
        
        for col in ['R', 'S', 'T', 'U', 'V', 'W']:
            cell = ws[f'{col}{row}']
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            self._apply_cell_style(cell)
        
        # Ajustar larguras
        for col in ['K', 'L', 'M', 'N', 'O', 'P', 'R', 'S', 'T', 'U', 'V', 'W']:
            ws.column_dimensions[col].width = 14
        
        # Coluna de separação mais estreita
        ws.column_dimensions['Q'].width = 3
    
    def _write_unpalletized_products(self, ws):
        """Escreve tabela de produtos não paletizados."""
        # Encontrar linha após a tabela de tipos (linha do último ALL + 3)
        # A tabela de tipos começa na linha 10, então calculamos dinamicamente
        start_row_types = 13  # Primeira linha de dados da tabela de tipos
        num_pallets = len(set(list(self.csharp_pallets.keys()) + list(self.python_pallets.keys())))
        row = start_row_types + num_pallets + 1 + 3  # +1 para linha de totais, +3 para espaçamento
        
        # Título
        ws[f'K{row}'] = 'Produtos Não Paletizados'
        ws[f'K{row}'].font = Font(bold=True, size=12, color="DC143C")
        ws.merge_cells(f'K{row}:W{row}')
        
        row += 2
        
        # Headers principais
        ws[f'K{row}'] = 'WMS'
        ws[f'K{row}'].font = Font(bold=True, size=11, color="366092")
        ws.merge_cells(f'K{row}:M{row}')
        ws[f'K{row}'].alignment = Alignment(horizontal="center", vertical="center")
        
        ws[f'O{row}'] = 'RESULTADO'
        ws[f'O{row}'].font = Font(bold=True, size=11, color="366092")
        ws.merge_cells(f'O{row}:Q{row}')
        ws[f'O{row}'].alignment = Alignment(horizontal="center", vertical="center")
        
        ws[f'S{row}'] = 'COMPARAÇÃO'
        ws[f'S{row}'].font = Font(bold=True, size=11, color="366092")
        ws.merge_cells(f'S{row}:U{row}')
        ws[f'S{row}'].alignment = Alignment(horizontal="center", vertical="center")
        
        row += 1
        
        # Headers colunas
        headers = [
            ('K', 'Código'), ('L', 'Nome'), ('M', 'Qtd'),
            ('N', ''),  # Separação
            ('O', 'Código'), ('P', 'Nome'), ('Q', 'Qtd'),
            ('R', ''),  # Separação
            ('S', 'Status'), ('T', 'Diff WMS'), ('U', 'Diff API')
        ]
        
        for col, header in headers:
            ws[f'{col}{row}'] = header
            if header:  # Não aplicar estilo em colunas de separação
                self._apply_header_style(ws[f'{col}{row}'])
        
        row += 1
        start_row = row
        
        # Combinar produtos de ambos os lados
        all_codes = set(list(self.csharp_outside.keys()) + list(self.python_outside.keys()))
        
        wms_list = []
        api_list = []
        
        for code in sorted(all_codes):
            wms_qty = self.csharp_outside.get(code, 0)
            api_qty = self.python_outside.get(code, 0)
            
            # Tentar obter nome do produto
            name = self._get_product_name_by_code(code)
            
            if wms_qty > 0:
                wms_list.append({'code': code, 'name': name, 'qty': wms_qty})
            if api_qty > 0:
                api_list.append({'code': code, 'name': name, 'qty': api_qty})
        
        # Escrever lado a lado
        max_rows = max(len(wms_list), len(api_list))
        
        for i in range(max_rows):
            # WMS
            if i < len(wms_list):
                ws[f'K{row}'] = wms_list[i]['code']
                ws[f'L{row}'] = wms_list[i]['name']
                ws[f'M{row}'] = wms_list[i]['qty']
                self._apply_cell_style(ws[f'K{row}'])
                self._apply_cell_style(ws[f'L{row}'])
                self._apply_cell_style(ws[f'M{row}'])
            
            # Separação
            ws[f'N{row}'] = ''
            
            # API
            if i < len(api_list):
                ws[f'O{row}'] = api_list[i]['code']
                ws[f'P{row}'] = api_list[i]['name']
                ws[f'Q{row}'] = api_list[i]['qty']
                self._apply_cell_style(ws[f'O{row}'])
                self._apply_cell_style(ws[f'P{row}'])
                self._apply_cell_style(ws[f'Q{row}'])
            
            # Separação
            ws[f'R{row}'] = ''
            
            # Comparação
            if i == 0:  # Apenas na primeira linha
                total_wms = sum(p['qty'] for p in wms_list)
                total_api = sum(p['qty'] for p in api_list)
                diff = total_api - total_wms
                
                if diff < 0:
                    ws[f'S{row}'] = 'API Melhor ✓'
                    bg_color = 'C6EFCE'
                elif diff == 0:
                    ws[f'S{row}'] = 'Igual'
                    bg_color = 'FFEB9C'
                else:
                    ws[f'S{row}'] = 'API Pior ✗'
                    bg_color = 'FFC7CE'
                
                ws[f'T{row}'] = total_wms
                ws[f'U{row}'] = total_api
                
                for col in ['S', 'T', 'U']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    ws[f'{col}{row}'].font = Font(bold=True)
                    self._apply_cell_style(ws[f'{col}{row}'])
            
            row += 1
        
        # Ajustar larguras (não alteramos N e R - mantém largura padrão)
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 40
        ws.column_dimensions['M'].width = 10
        ws.column_dimensions['O'].width = 12
        ws.column_dimensions['P'].width = 40
        ws.column_dimensions['Q'].width = 10
        ws.column_dimensions['S'].width = 15
        ws.column_dimensions['T'].width = 12
        ws.column_dimensions['U'].width = 12
    
    def _get_product_name_by_code(self, code: int) -> str:
        """Busca nome do produto pelo código nos pallets."""
        # Tentar em pallets WMS
        for pallet in self.csharp_pallets.values():
            for prod in pallet.products:
                if prod.code == code:
                    return prod.name
        
        # Tentar em pallets API
        for pallet in self.python_pallets.values():
            for prod in pallet.products:
                if prod.code == code:
                    return prod.name
        
        return f"Produto {code}"
    
    def _write_weight_comparison(self, ws):
        """Escreve comparação de pesos por lado."""
        # Encontrar linha após a tabela de não paletizados
        start_row_types = 13
        num_pallets = len(set(list(self.csharp_pallets.keys()) + list(self.python_pallets.keys())))
        num_outside = max(len(self.csharp_outside), len(self.python_outside))
        row_unpalletized = start_row_types + num_pallets + 1 + 3
        row = row_unpalletized + num_outside + 6  # +6 para título e headers da tabela anterior + 2 linhas
        
        # Título
        ws[f'K{row}'] = 'Comparação de Pesos por Lado'
        ws[f'K{row}'].font = Font(bold=True, size=12, color="366092")
        ws.merge_cells(f'K{row}:Q{row}')
        
        row += 2
        
        # Headers principais
        ws[f'K{row}'] = 'WMS'
        ws[f'K{row}'].font = Font(bold=True, size=11, color="366092")
        ws.merge_cells(f'K{row}:M{row}')
        ws[f'K{row}'].alignment = Alignment(horizontal="center", vertical="center")
        
        ws[f'O{row}'] = 'RESULTADO'
        ws[f'O{row}'].font = Font(bold=True, size=11, color="366092")
        ws.merge_cells(f'O{row}:Q{row}')
        ws[f'O{row}'].alignment = Alignment(horizontal="center", vertical="center")
        
        row += 1
        
        # Headers colunas
        ws[f'K{row}'] = 'Descrição'
        ws[f'L{row}'] = 'Peso'
        ws[f'M{row}'] = 'Percentual'
        ws[f'N{row}'] = ''  # Separação
        ws[f'O{row}'] = 'Descrição'
        ws[f'P{row}'] = 'Peso'
        ws[f'Q{row}'] = 'Percentual'
        
        for col in ['K', 'L', 'M', 'O', 'P', 'Q']:
            self._apply_header_style(ws[f'{col}{row}'])
        
        row += 1
        
        # Calcular pesos por lado
        wms_weights = {'M': 0.0, 'A': 0.0}
        api_weights = {'M': 0.0, 'A': 0.0}
        
        for pallet in self.csharp_pallets.values():
            wms_weights[pallet.side] += pallet.weight
        
        for pallet in self.python_pallets.values():
            api_weights[pallet.side] += pallet.weight
        
        wms_total = wms_weights['M'] + wms_weights['A']
        api_total = api_weights['M'] + api_weights['A']
        
        # WMS - Lado Motorista
        ws[f'K{row}'] = 'Lado Motorista'
        ws[f'L{row}'] = f"{wms_weights['M']:.2f}"
        ws[f'M{row}'] = f"{(wms_weights['M'] / wms_total * 100):.2f}%" if wms_total > 0 else "0.00%"
        
        # Separação
        ws[f'N{row}'] = ''
        
        # API - Lado Motorista
        ws[f'O{row}'] = 'Lado Motorista'
        ws[f'P{row}'] = f"{api_weights['M']:.2f}"
        ws[f'Q{row}'] = f"{(api_weights['M'] / api_total * 100):.2f}%" if api_total > 0 else "0.00%"
        
        for col in ['K', 'L', 'M', 'O', 'P', 'Q']:
            self._apply_cell_style(ws[f'{col}{row}'])
        
        row += 1
        
        # WMS - Lado Ajudante
        ws[f'K{row}'] = 'Lado Ajudante'
        ws[f'L{row}'] = f"{wms_weights['A']:.2f}"
        ws[f'M{row}'] = f"{(wms_weights['A'] / wms_total * 100):.2f}%" if wms_total > 0 else "0.00%"
        
        # Separação
        ws[f'N{row}'] = ''
        
        # API - Lado Ajudante
        ws[f'O{row}'] = 'Lado Ajudante'
        ws[f'P{row}'] = f"{api_weights['A']:.2f}"
        ws[f'Q{row}'] = f"{(api_weights['A'] / wms_total * 100):.2f}%" if api_total > 0 else "0.00%"
        
        for col in ['K', 'L', 'M', 'O', 'P', 'Q']:
            self._apply_cell_style(ws[f'{col}{row}'])
        
        row += 1
        
        # Linha de total
        ws[f'K{row}'] = 'TOTAL'
        ws[f'L{row}'] = f"{wms_total:.2f}"
        ws[f'M{row}'] = "100.00%"
        
        ws[f'N{row}'] = ''
        
        ws[f'O{row}'] = 'TOTAL'
        ws[f'P{row}'] = f"{api_total:.2f}"
        ws[f'Q{row}'] = "100.00%"
        
        for col in ['K', 'L', 'M', 'O', 'P', 'Q']:
            cell = ws[f'{col}{row}']
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            self._apply_cell_style(cell)
        
        # Ajustar larguras (não alteramos N - mantém largura padrão)
        ws.column_dimensions['K'].width = 18
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['O'].width = 18
        ws.column_dimensions['P'].width = 12
        ws.column_dimensions['Q'].width = 12
    
    def _guess_product_type(self, product_name: str) -> str:
        """Infere o tipo do produto baseado no nome."""
        name_upper = product_name.upper()
        
        # TopoPallet
        if any(keyword in name_upper for keyword in ['BAG IN BOX', 'MIX', 'RED BULL']):
            return 'TopoPallet'
        
        # Isotonico
        if 'GATORADE' in name_upper:
            return 'Isotonico'
        
        # Descartável (latas, PET, long neck, six-pack)
        if any(keyword in name_upper for keyword in ['LATA', 'PET', 'LONG NECK', 'SIX-PACK', 'SIXPACK', 'C/12', 'C12', 'C/24', 'SHRINK']):
            return 'Descartável'
        
        # Retornavel (garrafas, GFA, ML sem indicação de descartável)
        if any(keyword in name_upper for keyword in ['GFA', 'VD', 'CHOPP', 'PILSEN', 'ML']):
            return 'Retornavel'
        
        # Default: Descartável
        return 'Descartável'
    
    def _write_pivot_table(self, ws):
        """Escreve tabela pivot por pallet e produto."""
        row = 10
        
        # Título
        ws[f'B{row}'] = 'Tabela Pivot - Produtos por Pallet'
        ws[f'B{row}'].font = Font(bold=True, size=12, color="366092")
        
        row += 2
        
        # Headers
        ws[f'B{row}'] = 'Pallet ID'
        ws[f'C{row}'] = 'Código Produto'
        ws[f'D{row}'] = 'Nome Produto'
        ws[f'E{row}'] = 'Qtd WMS'
        ws[f'F{row}'] = 'Qtd API'
        ws[f'G{row}'] = 'Diferença'
        ws[f'H{row}'] = 'Status'
        
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            self._apply_header_style(ws[f'{col}{row}'])
        
        row += 1
        start_row = row
        
        # Coletar todos os pallets e produtos
        all_pallets = sorted(set(list(self.csharp_pallets.keys()) + list(self.python_pallets.keys())))
        
        for pallet_id in all_pallets:
            csharp_pallet = self.csharp_pallets.get(pallet_id)
            python_pallet = self.python_pallets.get(pallet_id)
            
            # Coletar produtos de ambos os lados
            products_map = {}
            
            if csharp_pallet:
                for prod in csharp_pallet.products:
                    products_map[prod.code] = {
                        'name': prod.name,
                        'csharp_qty': prod.quantity,
                        'python_qty': 0
                    }
            
            if python_pallet:
                for prod in python_pallet.products:
                    if prod.code in products_map:
                        products_map[prod.code]['python_qty'] = prod.quantity
                    else:
                        products_map[prod.code] = {
                            'name': prod.name,
                            'csharp_qty': 0,
                            'python_qty': prod.quantity
                        }
            
            # Escrever produtos
            for code in sorted(products_map.keys()):
                prod_info = products_map[code]
                
                ws[f'B{row}'] = pallet_id
                ws[f'C{row}'] = code
                ws[f'D{row}'] = prod_info['name']
                ws[f'E{row}'] = prod_info['csharp_qty']
                ws[f'F{row}'] = prod_info['python_qty']
                ws[f'G{row}'] = prod_info['python_qty'] - prod_info['csharp_qty']
                
                # Status
                if prod_info['csharp_qty'] == prod_info['python_qty']:
                    ws[f'H{row}'] = 'OK'
                    bg_color = 'C6EFCE'
                elif prod_info['csharp_qty'] == 0:
                    ws[f'H{row}'] = 'Novo API'
                    bg_color = 'FFEB9C'
                elif prod_info['python_qty'] == 0:
                    ws[f'H{row}'] = 'Falta API'
                    bg_color = 'FFC7CE'
                else:
                    ws[f'H{row}'] = 'Diferente'
                    bg_color = 'FFEB9C'
                
                for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    self._apply_cell_style(ws[f'{col}{row}'], bg_color if col == 'H' else None)
                
                row += 1
        
        # Ajustar larguras
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 15
    
    def _write_side_by_side(self, ws):
        """Escreve comparação lado a lado dos pallets."""
        # Encontrar próxima linha disponível
        row = ws.max_row + 3
        
        # Título
        ws[f'B{row}'] = 'Comparação Lado a Lado - Pallets'
        ws[f'B{row}'].font = Font(bold=True, size=12, color="366092")
        
        row += 2
        
        # Headers
        ws[f'B{row}'] = 'Pallet ID'
        ws[f'C{row}'] = 'Lado'
        ws[f'D{row}'] = 'Ocupação WMS'
        ws[f'E{row}'] = 'Ocupação API'
        ws[f'F{row}'] = 'Peso WMS'
        ws[f'G{row}'] = 'Peso API'
        ws[f'H{row}'] = 'Produtos WMS'
        ws[f'I{row}'] = 'Produtos API'
        ws[f'J{row}'] = 'Status'
        
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            self._apply_header_style(ws[f'{col}{row}'])
        
        row += 1
        
        # Coletar todos os pallets
        all_pallets = sorted(set(list(self.csharp_pallets.keys()) + list(self.python_pallets.keys())))
        
        for pallet_id in all_pallets:
            csharp_pallet = self.csharp_pallets.get(pallet_id)
            python_pallet = self.python_pallets.get(pallet_id)
            
            ws[f'B{row}'] = pallet_id
            ws[f'C{row}'] = csharp_pallet.side if csharp_pallet else (python_pallet.side if python_pallet else '-')
            ws[f'D{row}'] = f"{csharp_pallet.occupation:.2f}" if csharp_pallet else '-'
            ws[f'E{row}'] = f"{python_pallet.occupation:.2f}" if python_pallet else '-'
            ws[f'F{row}'] = f"{csharp_pallet.weight:.2f}" if csharp_pallet else '-'
            ws[f'G{row}'] = f"{python_pallet.weight:.2f}" if python_pallet else '-'
            ws[f'H{row}'] = len(csharp_pallet.products) if csharp_pallet else 0
            ws[f'I{row}'] = len(python_pallet.products) if python_pallet else 0
            
            # Status
            if csharp_pallet and python_pallet:
                if (abs(csharp_pallet.occupation - python_pallet.occupation) < 0.1 and 
                    len(csharp_pallet.products) == len(python_pallet.products)):
                    ws[f'J{row}'] = 'OK'
                    bg_color = 'C6EFCE'
                else:
                    ws[f'J{row}'] = 'Diferente'
                    bg_color = 'FFEB9C'
            elif not csharp_pallet:
                ws[f'J{row}'] = 'Novo API'
                bg_color = 'FFEB9C'
            else:
                ws[f'J{row}'] = 'Falta API'
                bg_color = 'FFC7CE'
            
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                self._apply_cell_style(ws[f'{col}{row}'], bg_color if col == 'J' else None)
            
            row += 1
        
        # Ajustar larguras
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 15
    
    def generate(self):
        """Gera o relatório Excel."""
        # Verificar se o arquivo já existe
        if self.output_file.exists():
            # Abrir arquivo existente
            wb = openpyxl.load_workbook(self.output_file)
            
            # Se a aba com o código da rota já existe, removê-la
            if self.route_code in wb.sheetnames:
                wb.remove(wb[self.route_code])
                print(f"Aba '{self.route_code}' existente removida.")
            
            # Criar nova aba
            ws = wb.create_sheet(title=self.route_code)
        else:
            # Criar novo arquivo
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.route_code
        
        # Escrever seções
        self._write_kpis(ws)
        self._write_pivot_table(ws)
        self._write_product_type_comparison(ws)
        self._write_unpalletized_products(ws)
        self._write_weight_comparison(ws)
        self._write_side_by_side(ws)
        
        # Salvar
        wb.save(self.output_file)
        print(f"\n{'='*80}")
        print(f"✓ Relatório gerado com sucesso!")
        print(f"✓ Aba '{self.route_code}' criada/atualizada")
        print(f"✓ Caminho completo: {self.output_file.absolute()}")
        print(f"{'='*80}\n")


def main():
    """Função principal para execução standalone."""
    import sys
    
    parser = argparse.ArgumentParser(description='Geracao de relatorio de comparacao')
    parser.add_argument('validation', nargs='?', default=r"c:\Users\BRKEY864393\OneDrive - Anheuser-Busch InBev\My Documents\projetos\ocp\ocp_score_ia\data\route\621387\output.txt", help='Path to output TXT')
    parser.add_argument('result', nargs='?', default=r"C:\Users\BRKEY864393\OneDrive - Anheuser-Busch InBev\My Documents\projetos\ocp\ocp_score_ia\data\route\621387\output\palletize_result_map_621387.txt", help='Path to output TXT')
    parser.add_argument('output', nargs='?', default=r"relatorio.xlsx", help='Output XLSX path (optional)')
    args = parser.parse_args()

    print("Iniciando...")
    
    # Se não foram passados argumentos, usa os defaults
    # Se foram passados argumentos mas menos que 2, mostra erro
    if len(sys.argv) > 1 and len(sys.argv) < 3:
        print("Uso: python gerar_relatorio.py <arquivo_csharp.txt> <arquivo_python.txt> [output.xlsx]")
        print("Ou execute sem argumentos para usar os caminhos padrão")
        sys.exit(1)
    
    validation_file = Path(args.validation)
    result_file = Path(args.result)
    output_file = Path(args.output)
    
    if not validation_file.exists():
        print(f"Arquivo não encontrado: {validation_file}")
        sys.exit(1)
    
    if not result_file.exists():
        print(f"Arquivo não encontrado: {result_file}")
        sys.exit(1)
    
    print("Gerando relatório...")
    print(f"validation_file: {validation_file}")
    print(f"result_file: {result_file}")
    print(f"output_file: {output_file}")
    generator = ReportGenerator(validation_file, result_file, output_file)
    generator.generate()


if __name__ == "__main__":
    main()
