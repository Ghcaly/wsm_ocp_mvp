import openpyxl

# Carregar o Excel
wb = openpyxl.load_workbook(r'c:\Users\BRKEY864393\OneDrive - Anheuser-Busch InBev\My Documents\projetos\ocp\ocp_score_ia\data\route\620807\output\comparacao_wms_vs_api.xlsx')
ws = wb['Comparação']

print('=== ANÁLISE DE SIMILARIDADE ===\n')

# Coletar todos os dados WMS
wms_data = []
for i in range(13, 100):
    mapa = ws[f'A{i}'].value
    if not mapa:
        break
    pallet = ws[f'B{i}'].value
    produto = ws[f'C{i}'].value
    qty = ws[f'E{i}'].value
    wms_data.append({'linha': i, 'mapa': mapa, 'pallet': pallet, 'produto': produto, 'qty': qty})

# Coletar todos os dados API
api_data = []
for i in range(13, 100):
    mapa = ws[f'I{i}'].value
    if not mapa:
        break
    pallet = ws[f'J{i}'].value
    produto = ws[f'K{i}'].value
    qty = ws[f'M{i}'].value
    api_data.append({'linha': i, 'mapa': mapa, 'pallet': pallet, 'produto': produto, 'qty': qty})

print(f'Total WMS: {len(wms_data)}')
print(f'Total API: {len(api_data)}\n')

# Criar set de API para comparação
api_set = set()
for item in api_data:
    key = (item['mapa'], item['pallet'], item['produto'], item['qty'])
    api_set.add(key)

# Verificar matches
matches = []
nao_matches = []

for wms in wms_data:
    key = (wms['mapa'], wms['pallet'], wms['produto'], wms['qty'])
    if key in api_set:
        matches.append(wms)
    else:
        nao_matches.append(wms)

print(f'✓ Linhas idênticas: {len(matches)} / {len(wms_data)} ({len(matches)/len(wms_data)*100:.1f}%)\n')

print('=== EXEMPLOS DE LINHAS QUE BATERAM (primeiras 5) ===')
for i, wms in enumerate(matches[:5]):
    print(f"{i+1}. Linha {wms['linha']}: mapa={wms['mapa']}, pallet={wms['pallet']}, produto={wms['produto']}, qty={wms['qty']}")

print('\n=== EXEMPLOS DE LINHAS QUE NÃO BATERAM (primeiras 10) ===')
for i, wms in enumerate(nao_matches[:10]):
    print(f"\n{i+1}. WMS Linha {wms['linha']}: mapa={wms['mapa']}, pallet={wms['pallet']}, produto={wms['produto']}, qty={wms['qty']}")
    
    # Procurar similar no API (mesmo pallet e produto)
    for api in api_data:
        if api['pallet'] == wms['pallet'] and api['produto'] == wms['produto']:
            print(f"   API Linha {api['linha']}: mapa={api['mapa']}, pallet={api['pallet']}, produto={api['produto']}, qty={api['qty']}")
            if api['mapa'] != wms['mapa']:
                print(f"   → DIFERENÇA: Mapa WMS={wms['mapa']} vs API={api['mapa']}")
            if api['qty'] != wms['qty']:
                print(f"   → DIFERENÇA: Qty WMS={wms['qty']} vs API={api['qty']}")
            break
    else:
        print(f"   → Produto {wms['produto']} não encontrado no pallet {wms['pallet']} no API")

wb.close()
